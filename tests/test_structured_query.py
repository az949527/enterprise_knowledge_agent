from __future__ import annotations

import asyncio
from pathlib import Path
import shutil
import tempfile
import unittest

from unittest.mock import AsyncMock, patch

from app.lite.computation_llm import (
    arun_computation_with_fallback,
    synthesize_mixed,
)
from app.lite.desktop_query import query_desktop_index
from app.lite.indexer import build_index, write_node_index
from app.lite.structured_query import (
    extract_computation_spec,
    run_structured_computation,
)
from app.security.remote_access import set_remote_access

_KEEP_ALIVE: list[tempfile.TemporaryDirectory] = []


def _department_budget_bytes() -> bytes:
    """程序化构造部门预算测试表（不依赖外部 fixture 数据）。"""
    from io import BytesIO

    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "预算"
    sheet.append(["部门", "第一季度", "第二季度", "第三季度", "第四季度"])
    sheet.append(["研发部", 95, 100, 110, 130])
    sheet.append(["销售部", 85, 90, 95, 105])
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _index_from_fixture(files: list[str]) -> Path:
    directory = tempfile.TemporaryDirectory()
    _KEEP_ALIVE.append(directory)
    source_dir = Path(directory.name) / "docs"
    source_dir.mkdir()
    content = _department_budget_bytes()
    for name in files:
        (source_dir / name).write_bytes(content)
    index_dir = Path(directory.name) / "index"
    build_index(source_dir, index_dir)
    return index_dir


async def _query(index_dir: Path, query: str) -> dict:
    return await query_desktop_index(
        query,
        index_dir,
        use_llm=False,
        llm_api_key="",
        llm_base_url="",
        llm_model="",
        use_embedding=False,
        use_reranker=False,
        retrieval_api_key="",
        offline=True,
    )


async def _query_with_history(
    index_dir: Path,
    query: str,
    history: list[dict[str, str]],
) -> dict:
    return await query_desktop_index(
        query,
        index_dir,
        use_llm=False,
        llm_api_key="",
        llm_base_url="",
        llm_model="",
        use_embedding=False,
        use_reranker=False,
        retrieval_api_key="",
        offline=True,
        conversation_history=history,
    )


def _build_multi_sheet_index() -> Path:
    """构造一个含 4 个 Sheet 的 xlsx 索引，用于多 Sheet 澄清场景。"""
    from io import BytesIO

    from openpyxl import Workbook

    directory = tempfile.TemporaryDirectory()
    _KEEP_ALIVE.append(directory)
    source_dir = Path(directory.name) / "docs"
    source_dir.mkdir()
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in ("Sheet1", "Sheet2", "Sheet3", "Sheet4"):
        sheet = workbook.create_sheet(name)
        sheet.append(["时间", "GDP", "CPI"])
        sheet.append(["2026-01", 12.5, 0.8])
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    (source_dir / "multi_sheet.xlsx").write_bytes(stream.getvalue())
    index_dir = Path(directory.name) / "index"
    build_index(source_dir, index_dir)
    return index_dir


def _build_duplicate_sheet_index() -> Path:
    """构造两个文件都有同名 Sheet1 的索引，用于多文件同名澄清场景。"""
    from io import BytesIO

    from openpyxl import Workbook

    directory = tempfile.TemporaryDirectory()
    _KEEP_ALIVE.append(directory)
    source_dir = Path(directory.name) / "docs"
    source_dir.mkdir()

    def _make(sheet_names: list[str], rows: int) -> bytes:
        workbook = Workbook()
        workbook.remove(workbook.active)
        for name in sheet_names:
            sheet = workbook.create_sheet(name)
            sheet.append(["时间", "GDP", "CPI"])
            for i in range(rows):
                sheet.append([f"2026-{i % 12 + 1:02d}", i, i + 0.5])
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        return stream.getvalue()

    (source_dir / "20230526.xlsx").write_bytes(_make(["Sheet1"], 196))
    (source_dir / "动态因子模型_data.xlsx").write_bytes(
        _make(["Sheet1", "Sheet2"], 122)
    )
    index_dir = Path(directory.name) / "index"
    build_index(source_dir, index_dir)
    return index_dir


class StructuredQueryTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.index_dir = _index_from_fixture(["budget.xlsx"])

    def _run(self, query: str):
        return run_structured_computation(query, self.index_dir)

    def test_sum(self) -> None:
        result = self._run("第四季度加起来是多少")
        self.assertEqual(result["mode"], "structured")
        self.assertIn("235", result["answer"])

    def test_avg(self) -> None:
        result = self._run("第一季度的平均值")
        self.assertEqual(result["mode"], "structured")
        self.assertIn("90", result["answer"])

    def test_argmax(self) -> None:
        result = self._run("哪个部门第四季度最高")
        self.assertEqual(result["mode"], "structured")
        self.assertIn("研发部", result["answer"])
        self.assertEqual(len(result["matched_rows"]), 1)
        self.assertIn("row_number", result["matched_rows"][0])

    def test_argmin(self) -> None:
        result = self._run("哪个部门第一季度最低")
        self.assertEqual(result["mode"], "structured")
        self.assertIn("销售部", result["answer"])

    def test_filter(self) -> None:
        result = self._run("第一季度超过90的部门")
        self.assertEqual(result["mode"], "structured")
        self.assertIn("研发部", result["answer"])

    def test_count(self) -> None:
        result = self._run("有多少行")
        self.assertEqual(result["mode"], "structured")
        self.assertIn("2 行", result["answer"])

    def test_sum_of_non_numeric_column_clarifies(self) -> None:
        result = self._run("部门加起来是多少")
        self.assertEqual(result["mode"], "structured_clarify")
        self.assertIn("数值", result["answer"])

    def test_filter_without_threshold_clarifies(self) -> None:
        result = self._run("第一季度的部门")
        # 无明确比较阈值 → 澄清或 fallback，不应误算
        self.assertIn(result["mode"], ("structured_clarify", "structured"))

    def test_row_evidence_carries_source(self) -> None:
        result = self._run("哪个部门第二季度最高")
        self.assertEqual(result["mode"], "structured")
        row = result["matched_rows"][0]
        self.assertIn("filename", row)
        self.assertIn("row_number", row)
        self.assertIn("cells", row)

    def test_multi_sheet_computation_clarifies_not_crash(self) -> None:
        """用户报告回归：多 Sheet Excel 走计算查询曾报
        'sequence item 0: expected str instance, set found'。"""
        from io import BytesIO

        from openpyxl import Workbook

        index_dir = _build_multi_sheet_index()
        result = run_structured_computation("文档里面有几行", index_dir)
        self.assertEqual(result["mode"], "structured_clarify")
        self.assertIn("Sheet", result["answer"])

    def test_clarification_followup_computes_specified_sheet(self) -> None:
        """用户报告回归：澄清追问后回答"sheet1"，应补全并计算该 Sheet 行数，
        而不是当作独立查询导致"资料不足"。"""
        index_dir = _build_multi_sheet_index()

        first = run_structured_computation("有多少行", index_dir)
        self.assertEqual(first["mode"], "structured_clarify")

        history = [
            {"role": "user", "content": "有多少行"},
            {"role": "assistant", "content": first["answer"]},
        ]
        result = asyncio.run(
            _query_with_history(index_dir, "sheet1", history)
        )
        self.assertEqual(result["mode"], "structured")
        self.assertIn("行", result["answer"])

    def test_multi_level_clarification_resolves_file_and_sheet(self) -> None:
        """用户报告回归：多文件同名 Sheet，指定 sheet1 后澄清文件，
        用户回答文件名简称"动态因子"，应结合历史计算该文件 Sheet1 行数。"""
        index_dir = _build_duplicate_sheet_index()

        history = [
            {"role": "user", "content": "sheet有几行"},
            {"role": "assistant", "content": "检测到文件内多个 Sheet（Sheet1、Sheet2），请指定要对哪个 Sheet 计算。"},
            {"role": "user", "content": "sheet1"},
            {"role": "assistant", "content": "多个文件都有 Sheet「Sheet1」（20230526.xlsx、动态因子模型_data.xlsx），请指定要对哪个文件计算。"},
        ]
        result = asyncio.run(_query_with_history(index_dir, "动态因子", history))
        self.assertEqual(result["mode"], "structured")
        self.assertIn("122 行", result["answer"])


class StructuredQueryCrossFileTests(unittest.TestCase):
    def test_cross_file_sum(self) -> None:
        index_dir = _index_from_fixture(["budget.xlsx", "budget2.xlsx"])
        result = run_structured_computation("所有 Excel 第四季度加起来是多少", index_dir)
        self.assertEqual(result["mode"], "structured")
        self.assertIn("470", result["answer"])
        files = {row["filename"] for row in result["matched_rows"]}
        self.assertEqual(files, {"budget.xlsx", "budget2.xlsx"})


class StructuredQueryIntegrationTests(unittest.TestCase):
    def test_query_desktop_index_routes_to_structured(self) -> None:
        index_dir = _index_from_fixture(["budget.xlsx"])
        result = asyncio.run(_query(index_dir, "第四季度加起来是多少"))
        self.assertEqual(result["mode"], "structured")
        self.assertIn("235", result["answer"])
        self.assertFalse(result["llm"]["enabled"])
        self.assertFalse(result["retrieval"]["remote"])
        self.assertIn("row_numbers", result["sources"][0])

    def test_mixed_index_routes_content_query_to_rag(self) -> None:
        directory = tempfile.TemporaryDirectory()
        _KEEP_ALIVE.append(directory)
        source_dir = Path(directory.name) / "docs"
        source_dir.mkdir()
        (source_dir / "budget.xlsx").write_bytes(_department_budget_bytes())
        (source_dir / "policy.md").write_text(
            "员工报销需在出差结束后三十天内提交。", encoding="utf-8"
        )
        index_dir = Path(directory.name) / "index"
        build_index(source_dir, index_dir)
        result = asyncio.run(_query(index_dir, "报销制度怎么规定"))
        self.assertEqual(result["mode"], "local_fallback")
        self.assertIn("报销", result["answer"])


class MixedComputationTests(unittest.TestCase):
    def test_mixed_computation_and_doc(self) -> None:
        directory = tempfile.TemporaryDirectory()
        _KEEP_ALIVE.append(directory)
        source_dir = Path(directory.name) / "docs"
        source_dir.mkdir()
        (source_dir / "budget.xlsx").write_bytes(_department_budget_bytes())
        (source_dir / "policy.md").write_text(
            "员工报销需在出差结束后三十天内提交。", encoding="utf-8"
        )
        index_dir = Path(directory.name) / "index"
        build_index(source_dir, index_dir)
        result = asyncio.run(
            _query(index_dir, "预算总和是多少以及报销制度怎么规定")
        )
        self.assertEqual(result["mode"], "mixed")
        self.assertIn("【计算结果】", result["answer"])
        self.assertIn("【相关资料】", result["answer"])

    def test_pure_computation_not_mixed(self) -> None:
        directory = tempfile.TemporaryDirectory()
        _KEEP_ALIVE.append(directory)
        source_dir = Path(directory.name) / "docs"
        source_dir.mkdir()
        (source_dir / "budget.xlsx").write_bytes(_department_budget_bytes())
        (source_dir / "policy.md").write_text(
            "员工报销需在出差结束后三十天内提交。", encoding="utf-8"
        )
        index_dir = Path(directory.name) / "index"
        build_index(source_dir, index_dir)
        result = asyncio.run(_query(index_dir, "第四季度加起来是多少"))
        self.assertEqual(result["mode"], "structured")


class ComputationFallbackTests(unittest.TestCase):
    def setUp(self) -> None:
        set_remote_access(True)

    def tearDown(self) -> None:
        set_remote_access(False)

    def test_no_key_fallback_not_triggered(self) -> None:
        index_dir = _index_from_fixture(["budget.xlsx"])
        result = asyncio.run(
            arun_computation_with_fallback(
                "部门加起来是多少", index_dir, llm_api_key=""
            )
        )
        self.assertEqual(result["mode"], "structured_clarify")

    @patch(
        "app.lite.computation_llm.resolve_column",
        new_callable=AsyncMock,
        return_value="第四季度",
    )
    def test_llm_resolves_ambiguous_column(self, mock_resolve) -> None:
        index_dir = _index_from_fixture(["budget.xlsx"])
        result = asyncio.run(
            arun_computation_with_fallback(
                "加起来是多少",
                index_dir,
                llm_api_key="sk-test",
                llm_base_url="https://example.test",
                llm_model="model-test",
            )
        )
        self.assertEqual(result["mode"], "structured")
        self.assertIn("235", result["answer"])
        mock_resolve.assert_awaited_once()

    @patch(
        "app.lite.computation_llm.resolve_column",
        new_callable=AsyncMock,
        return_value=None,
    )
    def test_llm_unable_to_resolve_keeps_clarify(self, mock_resolve) -> None:
        index_dir = _index_from_fixture(["budget.xlsx"])
        result = asyncio.run(
            arun_computation_with_fallback(
                "加起来是多少",
                index_dir,
                llm_api_key="sk-test",
                llm_base_url="https://example.test",
                llm_model="model-test",
            )
        )
        self.assertEqual(result["mode"], "structured_clarify")

    @patch(
        "app.lite.computation_llm._llm_call",
        new_callable=AsyncMock,
        return_value="综合答案",
    )
    def test_synthesize_mixed_uses_llm(self, mock_call) -> None:
        text = asyncio.run(
            synthesize_mixed(
                "预算总和和报销规定",
                "合计 = 235",
                [{"filename": "budget.xlsx", "row_numbers": [2], "content": "a"}],
                [{"filename": "policy.md", "chunk_index": 0, "content": "三十天内"}],
                api_key="sk-test",
                base_url="https://example.test",
                model="model-test",
            )
        )
        self.assertEqual(text, "综合答案")
        mock_call.assert_awaited_once()


class ChunkSplitReassemblyTests(unittest.TestCase):
    def test_long_row_group_uses_node_grid(self) -> None:
        # 生成行足够多、单行也长的 CSV，保证 split_text 会把 row_group 的 content 硬切。
        header = "部门,预算,负责人\n"
        rows = [
            f"部门{i},{10000 + i},负责人{i}" for i in range(80)
        ]
        directory = tempfile.TemporaryDirectory()
        _KEEP_ALIVE.append(directory)
        source_dir = Path(directory.name) / "docs"
        source_dir.mkdir()
        csv_path = source_dir / "big.csv"
        csv_path.write_text(header + "\n".join(rows) + "\n", encoding="utf-8")
        index_dir = Path(directory.name) / "index"
        build_index(source_dir, index_dir, chunk_size=120, chunk_overlap=20)
        result = run_structured_computation("预算加起来是多少", index_dir)
        self.assertEqual(result["mode"], "structured")
        # 80 行，预算 = 10000..10079，合计 = (10000+10079)*80/2 = 803_160
        self.assertIn("803160", result["answer"])


if __name__ == "__main__":
    unittest.main()
