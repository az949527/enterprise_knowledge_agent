from __future__ import annotations

import asyncio
import os
import sys
import time
from pathlib import Path
from typing import Any, Callable, Dict, List

from PySide6.QtCore import QSettings, QStandardPaths, Qt, QThread, QUrl, Signal
from PySide6.QtGui import (
    QColor,
    QCloseEvent,
    QDesktopServices,
    QImage,
    QKeySequence,
    QPainter,
    QPixmap,
    QShortcut,
)
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QDialog,
    QFileDialog,
    QFormLayout,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QListWidget,
    QMainWindow,
    QMessageBox,
    QPlainTextEdit,
    QProgressBar,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QSplitter,
    QStyle,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from app.lite.desktop_query import query_desktop_index
from app.lite.followup_rewriter import rewrite_followup
from app.lite.indexer import (
    IndexCancelledError,
    IndexFormatError,
    SUPPORTED_EXTENSIONS,
    delete_index_document,
    list_index_documents,
    sync_index_paths,
)
from app.lite.index_diagnostics import diagnose_index
from app.lite.query_planner import plan_query
from app.lite.remote_retrieval import (
    DEFAULT_EMBEDDING_MODEL,
    DEFAULT_RERANKER_MODEL,
    DEFAULT_RETRIEVAL_BASE_URL,
    set_remote_access,
)
from app.core.config import DEFAULT_LLM_MODEL, normalize_llm_model, settings

# 确保数据库路径稳定可用：打包 EXE 不依赖当前工作目录。
# engine 在 import database 时创建，必须在下面 import 之前设置 DATABASE_URL。
if getattr(sys, "frozen", False):
    _app_data = QStandardPaths.writableLocation(QStandardPaths.AppDataLocation)
    _db_dir = Path(_app_data).resolve() / "data"
    _db_dir.mkdir(parents=True, exist_ok=True)
    settings.DATABASE_URL = (
        "sqlite+aiosqlite:///"
        + (_db_dir / "enterprise_knowledge_agent.db").as_posix()
    )
else:
    Path("data").mkdir(parents=True, exist_ok=True)

from app.core.database import async_session_factory, init_db
from app.services.conversation_service import ConversationService
from app.security import (
    DEFAULT_SERVICE,
    get_backend_name,
    get_secret,
    set_secret,
)
from app.documents import (
    CSV_ENCODING_CANDIDATES,
    CsvEncodingError,
)


APP_NAME = "Local Knowledge Tool"
ORGANIZATION = "EnterpriseKnowledgeAgent"
SETTINGS_APP_NAME = "Local Knowledge Tool Desktop 1.0"
SETTINGS_SCHEMA_VERSION = 3
DESKTOP_FILE_FILTER = (
    "Knowledge files ("
    + " ".join(f"*{extension}" for extension in sorted(SUPPORTED_EXTENSIONS))
    + ")"
)


def desktop_index_dir() -> Path:
    override = os.getenv("DESKTOP_INDEX_DIR")
    if override:
        return Path(override).expanduser().resolve()
    if not getattr(sys, "frozen", False):
        return Path("data/lite_index").resolve()
    app_data = QStandardPaths.writableLocation(QStandardPaths.AppDataLocation)
    return Path(app_data).resolve() / "lite_index"


def desktop_temp_dir() -> Path:
    override = os.getenv("DESKTOP_TEMP_DIR")
    if override:
        return Path(override).expanduser().resolve()
    temp_root = QStandardPaths.writableLocation(QStandardPaths.TempLocation)
    return Path(temp_root).resolve() / "EnterpriseKnowledgeAgent"


def cleanup_stale_temp_files() -> None:
    """删除超过 7 天的应用临时残留文件。PDF 页面图片不会落盘。"""
    root = desktop_temp_dir()
    if not root.exists():
        return
    cutoff = time.time() - 7 * 24 * 3600
    for path in root.iterdir():
        try:
            if path.is_file() and path.stat().st_mtime < cutoff:
                path.unlink()
        except OSError:
            continue


def cleanup_temp_files() -> None:
    """正常退出时清理本进程创建的应用临时文件。"""
    root = desktop_temp_dir()
    if not root.exists():
        return
    try:
        for path in root.iterdir():
            try:
                path.unlink()
            except OSError:
                continue
        root.rmdir()
    except OSError:
        pass


class IndexWorker(QThread):
    completed = Signal(dict)
    progress_changed = Signal(dict)
    encoding_required = Signal(str, str)
    cancelled = Signal()
    failed = Signal(str)

    def __init__(
        self,
        paths: List[Path],
        index_dir: Path,
        csv_encodings: Dict[str, str] | None = None,
        source_root: Path | None = None,
        force_reparse: bool = False,
        replace_all: bool = False,
    ) -> None:
        super().__init__()
        self.paths = paths
        self.index_dir = index_dir
        self.csv_encodings = dict(csv_encodings or {})
        self.source_root = source_root
        self.force_reparse = force_reparse
        self.replace_all = replace_all

    def run(self) -> None:
        try:
            stats = sync_index_paths(
                self.paths,
                self.index_dir,
                source_root=self.source_root,
                source_label=(
                    self.source_root.as_posix()
                    if self.source_root is not None
                    else "desktop_upload"
                ),
                remove_missing=self.replace_all,
                force_reparse=self.force_reparse,
                csv_encodings=self.csv_encodings,
                progress=self.progress_changed.emit,
                should_cancel=self.isInterruptionRequested,
            )
            self.completed.emit(stats.__dict__)
        except CsvEncodingError as exc:
            failed_path = next(
                (
                    path
                    for path in self.paths
                    if path.name == Path(exc.source_path).name
                ),
                Path(exc.source_path),
            )
            self.encoding_required.emit(str(failed_path.resolve()), str(exc))
        except IndexCancelledError:
            self.cancelled.emit()
        except Exception as exc:
            self.failed.emit(str(exc))


class DiagnosticWorker(QThread):
    completed = Signal(dict)
    failed = Signal(str)

    def __init__(self, index_dir: Path) -> None:
        super().__init__()
        self.index_dir = index_dir

    def run(self) -> None:
        try:
            self.completed.emit(diagnose_index(self.index_dir))
        except Exception as exc:
            self.failed.emit(str(exc))


class QueryWorker(QThread):
    """在后台线程完成整条查询链路，避免阻塞 UI 线程。

    覆盖：会话准备(创建/选择) → 读取历史 → 追问改写(可能远程)
        → 保存用户消息 → 检索/生成 → 保存助手消息 → 自动摘要。
    """
    completed = Signal(dict)
    failed = Signal(str)

    def __init__(
        self,
        query: str,
        index_dir: Path,
        *,
        conv_id: str | None,
        use_llm: bool,
        llm_api_key: str,
        llm_base_url: str,
        llm_model: str,
        use_embedding: bool,
        use_reranker: bool,
        retrieval_api_key: str,
        retrieval_base_url: str,
        embedding_model: str,
        reranker_model: str,
        offline: bool = False,
    ) -> None:
        super().__init__()
        self.query = query
        self.index_dir = index_dir
        self.conv_id = conv_id
        self.use_llm = use_llm
        self.llm_api_key = llm_api_key
        self.llm_base_url = llm_base_url
        self.llm_model = llm_model
        self.use_embedding = use_embedding
        self.use_reranker = use_reranker
        self.retrieval_api_key = retrieval_api_key
        self.retrieval_base_url = retrieval_base_url
        self.embedding_model = embedding_model
        self.reranker_model = reranker_model
        self.offline = offline

    def run(self) -> None:
        self._resolved_conv_id = self.conv_id
        try:
            result = asyncio.run(self._run())
            self.completed.emit(result)
        except Exception as exc:
            # 后台保存错误消息（UI 线程不阻塞），便于用户回溯
            try:
                asyncio.run(
                    _save_error_message(self._resolved_conv_id, str(exc))
                )
            except Exception:
                pass
            self.failed.emit(str(exc))

    async def _run(self) -> dict:
        # 1. 会话准备：无会话则自动创建
        conv_id = self.conv_id
        async with async_session_factory() as db:
            if not conv_id:
                conv = await ConversationService.create_conversation(db)
                await db.commit()
                conv_id = conv.id
            self._resolved_conv_id = conv_id
            recent_msgs = await ConversationService.get_messages(
                db, conv_id, limit=10
            )
        history = []
        for message in recent_msgs[-6:]:
            if message.role == "user" and message.original_query:
                history.append({"role": "user", "content": message.original_query})
            elif message.role == "assistant" and message.answer:
                history.append({"role": "assistant", "content": message.answer})

        # 2. 追问改写（可能发起远程请求，放在后台线程）
        rewritten_query = self.query
        if history:
            try:
                rewritten_query = await rewrite_followup(
                    self.query,
                    history,
                    api_key=self.llm_api_key,
                    base_url=self.llm_base_url,
                    model=self.llm_model,
                )
            except Exception:
                rewritten_query = self.query

        # 3. 保存用户消息
        async with async_session_factory() as db:
            await ConversationService.add_message(
                db,
                conv_id,
                role="user",
                original_query=self.query,
                rewritten_query=(
                    rewritten_query if rewritten_query != self.query else None
                ),
            )
            await db.commit()

        # 4. 检索与生成
        result = await query_desktop_index(
            rewritten_query,
            self.index_dir,
            top_k=5,
            use_llm=self.use_llm,
            llm_api_key=self.llm_api_key,
            llm_base_url=self.llm_base_url,
            llm_model=self.llm_model,
            use_embedding=self.use_embedding,
            use_reranker=self.use_reranker,
            retrieval_api_key=self.retrieval_api_key,
            retrieval_base_url=self.retrieval_base_url,
            embedding_model=self.embedding_model,
            reranker_model=self.reranker_model,
            offline=self.offline,
            conversation_history=history or None,
        )

        # 5. 保存助手消息
        llm_info = result.get("llm") or {}
        mode = str(result.get("mode") or "")
        async with async_session_factory() as db:
            await ConversationService.add_message(
                db,
                conv_id,
                role="assistant",
                answer=str(result.get("answer") or ""),
                citations=result.get("sources") or [],
                model=llm_info.get("model") or llm_info.get("configured_model") or "",
                token_usage=llm_info.get("usage"),
                error=llm_info.get("error") if mode == "llm_error" else None,
            )
            await db.commit()

        # 6. 自动摘要（后台）
        try:
            await _maybe_summarize_conversation(
                conv_id,
                api_key=self.llm_api_key,
                base_url=self.llm_base_url,
                model=self.llm_model,
                offline=self.offline,
            )
        except Exception:
            pass

        result["conv_id"] = conv_id
        result["original_query"] = self.query
        result["rewritten_query"] = (
            rewritten_query if rewritten_query != self.query else ""
        )
        return result


def _desktop_db_path() -> Path | None:
    """从 settings.DATABASE_URL 解析本地 SQLite 文件路径。"""
    url = str(settings.DATABASE_URL or "")
    if not url.startswith("sqlite"):
        return None
    for prefix in ("sqlite+aiosqlite:///", "sqlite:///", "sqlite://"):
        if url.startswith(prefix):
            raw = url[len(prefix):]
            if raw:
                return Path(raw).expanduser().resolve()
    return None


def _human_size(num_bytes: int) -> str:
    value = float(num_bytes)
    for unit in ("B", "KB", "MB", "GB"):
        if value < 1024 or unit == "GB":
            return f"{int(value)} {unit}" if unit == "B" else f"{value:.1f} {unit}"
        value /= 1024
    return f"{value:.1f} GB"


def _dir_size(path: Path) -> int:
    """递归统计目录/文件字节数，不跟随符号链接，忽略权限错误。"""
    total = 0
    try:
        if path.is_file():
            try:
                total += path.stat().st_size
            except OSError:
                return 0
        elif path.is_dir():
            for item in path.rglob("*"):
                if item.is_file():
                    try:
                        total += item.stat().st_size
                    except OSError:
                        continue
    except OSError:
        return 0
    return total


class DiskStatsWorker(QThread):
    """后台统计索引与聊天数据库占用，避免阻塞 UI。"""
    completed = Signal(dict)

    def __init__(self, index_dir: Path, db_path: Path | None) -> None:
        super().__init__()
        self.index_dir = index_dir
        self.db_path = db_path

    def run(self) -> None:
        index_bytes = _dir_size(self.index_dir)
        db_bytes = 0
        if self.db_path is not None:
            for candidate in (self.db_path,):
                db_bytes += _dir_size(candidate)
            for suffix in ("-wal", "-shm", "-journal"):
                sidecar = self.db_path.with_name(self.db_path.name + suffix)
                db_bytes += _dir_size(sidecar)
        self.completed.emit({"index_bytes": index_bytes, "db_bytes": db_bytes})


class _PdfPageWorker(QThread):
    """后台渲染 PDF 指定页并叠加引用高亮框。"""
    loaded = Signal(object, int, int)  # QImage, page_number(1-based), page_count
    failed = Signal(str)

    def __init__(self, origin_path: Path, page_number: int, zoom: float,
                 highlight_bbox: list | None) -> None:
        super().__init__()
        self.origin_path = origin_path
        self.page_number = page_number
        self.zoom = zoom
        self.highlight_bbox = highlight_bbox

    def run(self) -> None:
        try:
            import fitz

            with fitz.open(self.origin_path) as document:
                page = document.load_page(self.page_number)
                matrix = fitz.Matrix(self.zoom, self.zoom)
                pix = page.get_pixmap(matrix=matrix, alpha=False)
                image = QImage(
                    pix.samples,
                    pix.width,
                    pix.height,
                    pix.stride,
                    QImage.Format_RGB888,
                ).copy()
                if self.highlight_bbox and len(self.highlight_bbox) == 4:
                    x0, y0, x1, y1 = self.highlight_bbox
                    painter = QPainter(image)
                    painter.setPen(QColor(230, 60, 60, 220))
                    painter.drawRect(
                        int(x0 * self.zoom),
                        int(y0 * self.zoom),
                        int((x1 - x0) * self.zoom),
                        int((y1 - y0) * self.zoom),
                    )
                    painter.end()
                self.loaded.emit(image, self.page_number + 1, document.page_count)
        except Exception as exc:
            self.failed.emit(str(exc))


class _ExcelSheetWorker(QThread):
    """后台加载 Excel 指定 Sheet 的数据用于预览。"""
    loaded = Signal(object, object, object)  # headers, rows, highlight(row_start,row_end)
    failed = Signal(str)

    def __init__(self, origin_path: Path, sheet_name: str,
                 row_start: int | None, row_end: int | None) -> None:
        super().__init__()
        self.origin_path = origin_path
        self.sheet_name = sheet_name
        self.row_start = row_start
        self.row_end = row_end

    def run(self) -> None:
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(
                self.origin_path, read_only=True, data_only=True
            )
            try:
                worksheet = (
                    workbook[self.sheet_name]
                    if self.sheet_name and self.sheet_name in workbook.sheetnames
                    else workbook.active
                )
                headers = []
                rows = []
                for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
                    values = ["" if value is None else str(value) for value in row]
                    if row_index == 0:
                        headers = values
                    else:
                        rows.append(values)
            finally:
                workbook.close()
            self.loaded.emit(headers, rows, (self.row_start, self.row_end))
        except Exception as exc:
            self.failed.emit(str(exc))


class SourcePreviewDialog(QDialog):
    """应用内引用定位预览：PDF 渲染页 / Excel 表格，原文件缺失时退化为索引内容。"""

    def __init__(
        self,
        source: Dict[str, Any],
        origin_path: Path,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.source = source
        self.origin_path = origin_path
        self.setWindowTitle("引用定位预览")
        self.resize(760, 560)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        title = QLabel(str(source.get("filename") or "未知文件"))
        title.setObjectName("pageTitle")
        layout.addWidget(title)

        if not origin_path.exists():
            self._build_index_fallback(layout, source)
            return

        suffix = origin_path.suffix.casefold()
        if suffix == ".pdf":
            self._build_pdf_preview(layout, source)
        elif suffix in (".xlsx", ".xls"):
            self._build_excel_preview(layout, source)
        else:
            self._build_index_fallback(layout, source)

    # ---------- PDF ----------
    def _build_pdf_preview(self, layout: QVBoxLayout, source: Dict[str, Any]) -> None:
        anchor = source.get("source_anchor") or {}
        page_number = int(anchor.get("page") or source.get("page_or_sheet") or 1) - 1
        page_number = max(page_number, 0)
        highlight = source.get("bbox")
        self._pdf_page = page_number
        self._pdf_zoom = 1.5

        self._pdf_label = QLabel("加载中…")
        self._pdf_label.setAlignment(Qt.AlignCenter)
        self._pdf_label.setMinimumHeight(420)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(self._pdf_label)
        layout.addWidget(scroll, 1)

        controls = QHBoxLayout()
        prev_btn = QPushButton("上一页")
        prev_btn.clicked.connect(lambda: self._pdf_render(self._pdf_page - 1))
        controls.addWidget(prev_btn)
        self._pdf_page_label = QLabel("")
        self._pdf_page_label.setObjectName("mutedText")
        controls.addWidget(self._pdf_page_label)
        next_btn = QPushButton("下一页")
        next_btn.clicked.connect(lambda: self._pdf_render(self._pdf_page + 1))
        controls.addWidget(next_btn)
        controls.addStretch()
        zoom_out = QPushButton("缩小")
        zoom_out.clicked.connect(lambda: self._set_pdf_zoom(self._pdf_zoom * 0.8))
        controls.addWidget(zoom_out)
        zoom_in = QPushButton("放大")
        zoom_in.clicked.connect(lambda: self._set_pdf_zoom(self._pdf_zoom * 1.25))
        controls.addWidget(zoom_in)
        open_btn = QPushButton("打开原文件")
        open_btn.clicked.connect(
            lambda: QDesktopServices.openUrl(QUrl.fromLocalFile(str(self.origin_path)))
        )
        controls.addWidget(open_btn)
        layout.addLayout(controls)

        self._pdf_render(self._pdf_page)

    def _pdf_render(self, page_number: int) -> None:
        page_count = getattr(self, "_pdf_page_count", None)
        if page_count is not None and page_number < 0:
            return
        if page_count is not None and page_number >= page_count:
            return
        self._pdf_page = page_number
        worker = _PdfPageWorker(
            self.origin_path,
            max(page_number, 0),
            self._pdf_zoom,
            self.source.get("bbox"),
        )
        worker.loaded.connect(self._pdf_loaded)
        worker.failed.connect(lambda msg: self._pdf_label.setText(f"加载失败：{msg}"))
        worker.finished.connect(lambda: worker.deleteLater())
        worker.start()

    def _pdf_loaded(self, image: QImage, page_number: int, page_count: int) -> None:
        self._pdf_page_count = page_count
        self._pdf_label.setPixmap(QPixmap.fromImage(image))
        self._pdf_page_label.setText(f"{page_number} / {page_count}")

    def _set_pdf_zoom(self, zoom: float) -> None:
        zoom = min(max(zoom, 0.5), 5.0)
        self._pdf_zoom = zoom
        self._pdf_render(self._pdf_page)

    # ---------- Excel ----------
    def _build_excel_preview(self, layout: QVBoxLayout, source: Dict[str, Any]) -> None:
        anchor = source.get("source_anchor") or {}
        sheet_name = str(anchor.get("sheet") or source.get("page_or_sheet") or "")
        row_start = anchor.get("row_start") or source.get("row_start")
        row_end = anchor.get("row_end") or source.get("row_end")
        self._excel_highlight = (row_start, row_end)

        self._table = QTableWidget(0, 0)
        self._table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._table.verticalHeader().setVisible(False)
        layout.addWidget(self._table, 1)

        controls = QHBoxLayout()
        self._excel_info = QLabel("加载中…")
        self._excel_info.setObjectName("mutedText")
        controls.addWidget(self._excel_info)
        controls.addStretch()
        open_btn = QPushButton("打开原文件")
        open_btn.clicked.connect(
            lambda: QDesktopServices.openUrl(QUrl.fromLocalFile(str(self.origin_path)))
        )
        controls.addWidget(open_btn)
        layout.addLayout(controls)

        worker = _ExcelSheetWorker(self.origin_path, sheet_name, row_start, row_end)
        worker.loaded.connect(self._excel_loaded)
        worker.failed.connect(lambda msg: self._excel_info.setText(f"加载失败：{msg}"))
        worker.finished.connect(lambda: worker.deleteLater())
        worker.start()

    def _excel_loaded(self, headers: list, rows: list, highlight: tuple) -> None:
        self._table.setColumnCount(len(headers) or 1)
        self._table.setHorizontalHeaderLabels(
            [str(h) if h else "" for h in headers] or [""]
        )
        self._table.setRowCount(len(rows))
        for row_index, row in enumerate(rows):
            for col_index in range(len(headers) or len(row)):
                value = row[col_index] if col_index < len(row) else ""
                item = QTableWidgetItem(str(value))
                self._table.setItem(row_index, col_index, item)
        row_start, row_end = highlight
        if row_start is not None and row_end is not None:
            start = max(int(row_start) - 1, 0)
            end = min(int(row_end) - 1, len(rows) - 1)
            if start <= end:
                self._table.setRangeSelected(
                    self._table.model().index(start, 0),
                    self._table.model().index(end, self._table.columnCount() - 1),
                )
                self._table.scrollToItem(self._table.item(start, 0))
        self._excel_info.setText(
            f"{len(rows)} 行 × {len(headers)} 列"
            + (f"，高亮行 {row_start}-{row_end}" if row_start is not None else "")
        )

    # ---------- 索引内容兜底 ----------
    def _build_index_fallback(self, layout: QVBoxLayout, source: Dict[str, Any]) -> None:
        note = QLabel("原文件不存在，显示索引中保存的内容：")
        note.setObjectName("mutedText")
        layout.addWidget(note)
        content = QPlainTextEdit()
        content.setReadOnly(True)
        content.setPlainText(
            str(source.get("display_content") or source.get("content") or "")
        )
        layout.addWidget(content, 1)


SUMMARY_PROMPT = """你是一个对话摘要助手。给定一段多轮问答记录，请用不超过150字的中文概括：

1. 用户主要关心的问题或话题
2. 已获得的关键结论或数据（如果有）

只输出摘要文本，不要加前缀或解释。

对话记录：
{transcript}

摘要："""


async def _summarize_with_llm(
    transcript: str,
    api_key: str,
    base_url: str,
    model: str,
) -> str:
    """调用 LLM 生成对话摘要。失败时返回空字符串，由调用方回退规则提取。"""
    try:
        from openai import AsyncOpenAI

        client = AsyncOpenAI(api_key=api_key, base_url=base_url)
        response = await client.chat.completions.create(
            model=model,
            messages=[
                {
                    "role": "user",
                    "content": SUMMARY_PROMPT.format(transcript=transcript),
                }
            ],
            temperature=0.3,
            timeout=15,
        )
        text = (response.choices[0].message.content or "").strip()
        return text
    except Exception:
        return ""


async def _save_error_message(conv_id: str | None, message: str) -> None:
    """后台保存一条助手错误消息，避免 UI 线程阻塞。"""
    if not conv_id:
        return
    async with async_session_factory() as db:
        await ConversationService.add_message(
            db, conv_id, role="assistant", error=message
        )
        await db.commit()


async def _maybe_summarize_conversation(
    conv_id: str,
    *,
    api_key: str = "",
    base_url: str = "",
    model: str = "",
    offline: bool = True,
) -> None:
    """当消息数超过阈值时生成对话摘要。在后台线程调用，不阻塞 UI。"""
    if not conv_id:
        return
    async with async_session_factory() as db:
        conv = await ConversationService.get_conversation(db, conv_id)
        if (
            conv is None
            or (conv.message_count or 0) < settings.SUMMARY_TRIGGER_MESSAGE_COUNT
        ):
            return
        summary_text, recent_msgs = await ConversationService.get_active_context(
            db, conv_id
        )
        if summary_text and len(recent_msgs) <= 4:
            return
        transcript_lines = []
        user_questions = []
        for message in recent_msgs[-12:]:
            if message.role == "user" and message.original_query:
                transcript_lines.append(f"[用户] {message.original_query}")
                user_questions.append(message.original_query)
            elif message.role == "assistant":
                answer = message.answer or message.error or "(无回答)"
                transcript_lines.append(f"[助手] {answer[:200]}")
        transcript = "\n".join(transcript_lines)
        if not user_questions:
            return
        llm_summary = ""
        if api_key and not offline:
            llm_summary = await _summarize_with_llm(
                transcript, api_key, base_url, model
            )
        if not llm_summary:
            topics = "；".join(question[:60] for question in user_questions[:4])
            llm_summary = f"对话涉及以下话题：{topics}"
        if recent_msgs:
            await ConversationService.generate_summary(
                db,
                conv_id,
                llm_summary_text=llm_summary,
                start_message_id=recent_msgs[0].id,
                end_message_id=recent_msgs[-1].id,
                token_count=len(llm_summary) // 2,
            )
        await db.commit()


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.index_dir = desktop_index_dir()
        self.index_dir.mkdir(parents=True, exist_ok=True)
        self.settings = QSettings(ORGANIZATION, SETTINGS_APP_NAME)
        self.settings.setFallbacksEnabled(False)
        cleanup_stale_temp_files()
        self._initialize_release_settings()
        self._workers: List[QThread] = []
        self._active_index_worker: IndexWorker | None = None
        self._settings_loading = False
        self._network_active = False
        self._current_conv_id: str | None = None
        self._conversations: list = []

        self.setWindowTitle("本地知识库")
        self.setMinimumSize(780, 620)
        self.resize(980, 760)
        self.setStyleSheet(APP_STYLE)

        tabs = QTabWidget()
        tabs.setDocumentMode(True)
        tabs.addTab(self._build_chat_tab(), "对话")
        tabs.addTab(self._build_knowledge_tab(), "知识库")
        tabs.addTab(self._build_settings_tab(), "设置")
        self.setCentralWidget(tabs)

        self.statusBar().showMessage("就绪")
        self._load_settings()
        self.refresh_documents()
        self._ensure_chat_db()
        self._load_conversations()
        self._cleanup_old_data()
        self.start_disk_stats()

    def _initialize_release_settings(self) -> None:
        current_version = int(self.settings.value("release/schema_version", 0) or 0)
        if current_version >= SETTINGS_SCHEMA_VERSION:
            return
        if current_version < 1:
            self.settings.remove("llm/api_key")
        if current_version < 2:
            self.settings.remove("retrieval/api_key")
        if current_version < 3:
            # 迁移旧明文 API Key：读入系统凭据库后从设置中删除。
            legacy_llm = self.settings.value("llm/api_key", "", str)
            legacy_retrieval = self.settings.value("retrieval/api_key", "", str)
            if legacy_llm:
                set_secret(DEFAULT_SERVICE, "llm_api_key", legacy_llm)
            if legacy_retrieval:
                set_secret(DEFAULT_SERVICE, "retrieval_api_key", legacy_retrieval)
            self.settings.remove("llm/api_key")
            self.settings.remove("retrieval/api_key")
        self.settings.setValue("release/schema_version", SETTINGS_SCHEMA_VERSION)
        self.settings.sync()

    def _build_chat_tab(self) -> QWidget:
        page = QWidget()
        outer = QSplitter(Qt.Horizontal)
        outer_layout = QVBoxLayout(page)
        outer_layout.setContentsMargins(0, 0, 0, 0)
        outer_layout.addWidget(outer)

        # ---- Left: Conversation sidebar ----
        sidebar = QWidget()
        sidebar.setMaximumWidth(220)
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(12, 12, 6, 12)
        sidebar_layout.setSpacing(6)

        conv_heading = QLabel("对话列表")
        conv_heading.setObjectName("sectionTitle")
        sidebar_layout.addWidget(conv_heading)

        self.conv_search_input = QLineEdit()
        self.conv_search_input.setPlaceholderText("搜索对话...")
        self.conv_search_input.textChanged.connect(self._filter_conversations)
        sidebar_layout.addWidget(self.conv_search_input)

        self.conv_new_btn = QPushButton("+ 新对话")
        self.conv_new_btn.clicked.connect(self._new_conversation)
        sidebar_layout.addWidget(self.conv_new_btn)

        self.conv_list = QListWidget()
        self.conv_list.currentRowChanged.connect(self._conversation_selected)
        sidebar_layout.addWidget(self.conv_list, 1)

        conv_btn_row = QHBoxLayout()
        self.conv_delete_btn = QPushButton("删除")
        self.conv_delete_btn.setObjectName("dangerButton")
        self.conv_delete_btn.clicked.connect(self._delete_conversation)
        conv_btn_row.addWidget(self.conv_delete_btn)
        self.conv_export_btn = QPushButton("导出")
        self.conv_export_btn.clicked.connect(self._export_conversation)
        conv_btn_row.addWidget(self.conv_export_btn)
        sidebar_layout.addLayout(conv_btn_row)

        self.conv_clear_all_btn = QPushButton("清空全部")
        self.conv_clear_all_btn.setObjectName("dangerButton")
        self.conv_clear_all_btn.clicked.connect(self._clear_all_conversations)
        sidebar_layout.addWidget(self.conv_clear_all_btn)

        outer.addWidget(sidebar)

        # ---- Right: Chat area ----
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(6, 12, 12, 12)
        right_layout.setSpacing(10)

        heading = QLabel("知识库问答")
        heading.setObjectName("pageTitle")
        right_layout.addWidget(heading)

        # Message history
        self.message_edit = QPlainTextEdit()
        self.message_edit.setReadOnly(True)
        self.message_edit.setPlaceholderText("选择或创建对话开始提问")
        self.message_edit.setMaximumBlockCount(5000)
        right_layout.addWidget(self.message_edit, 1)

        # Input area
        self.query_input = QPlainTextEdit()
        self.query_input.setPlaceholderText("输入与已添加文档相关的问题 (Ctrl+Enter 发送)")
        self.query_input.setMaximumHeight(80)
        QShortcut(QKeySequence("Ctrl+Return"), self.query_input, self.ask_question)
        right_layout.addWidget(self.query_input)

        actions = QHBoxLayout()
        self.use_llm_checkbox = QCheckBox("使用 LLM 汇总答案")
        self.use_llm_checkbox.setChecked(True)
        actions.addWidget(self.use_llm_checkbox)
        self.use_embedding_checkbox = QCheckBox("远程 Embedding")
        actions.addWidget(self.use_embedding_checkbox)
        self.use_reranker_checkbox = QCheckBox("远程 Reranker")
        actions.addWidget(self.use_reranker_checkbox)
        actions.addStretch()

        self.network_status_label = QLabel("模式：—")
        self.network_status_label.setObjectName("networkStatus")
        actions.addWidget(self.network_status_label)

        self.ask_button = QPushButton("查询")
        self.ask_button.setIcon(self.style().standardIcon(QStyle.SP_ArrowForward))
        self.ask_button.clicked.connect(self.ask_question)
        actions.addWidget(self.ask_button)
        right_layout.addLayout(actions)

        # Source panel (collapsible)
        sources_label = QLabel("引用来源")
        sources_label.setObjectName("sectionTitle")
        right_layout.addWidget(sources_label)

        self.sources_container = QWidget()
        self.sources_layout = QVBoxLayout(self.sources_container)
        self.sources_layout.setContentsMargins(0, 0, 0, 0)
        self.sources_layout.setSpacing(8)
        self.sources_layout.addStretch()

        sources_scroll = QScrollArea()
        sources_scroll.setWidgetResizable(True)
        sources_scroll.setFrameShape(QFrame.NoFrame)
        sources_scroll.setMaximumHeight(200)
        sources_scroll.setWidget(self.sources_container)
        right_layout.addWidget(sources_scroll)

        outer.addWidget(right_panel)
        outer.setSizes([200, 760])
        return page

    # ==================== P1-3 Conversation Management ====================

    def _ensure_chat_db(self) -> None:
        """确保聊天相关数据表已创建。"""
        try:
            asyncio.run(init_db())
        except Exception:
            pass

    def _load_conversations(self) -> None:
        """从数据库加载会话列表到侧边栏。"""
        async def _load():
            async with async_session_factory() as db:
                return await ConversationService.list_conversations(db)

        try:
            self._conversations = asyncio.run(_load())
        except Exception:
            self._conversations = []

        current_row = self.conv_list.currentRow()
        self.conv_list.blockSignals(True)
        self.conv_list.clear()
        for conv in self._conversations:
            self.conv_list.addItem(
                f"{conv.title}  ({conv.message_count})"
            )
        if 0 <= current_row < len(self._conversations):
            self.conv_list.setCurrentRow(current_row)
        elif self._conversations and self._current_conv_id:
            # Re-select
            for i, c in enumerate(self._conversations):
                if c.id == self._current_conv_id:
                    self.conv_list.setCurrentRow(i)
                    break
        self.conv_list.blockSignals(False)

    def _set_chat_text(self, text: str) -> None:
        """设置聊天框文本并滚动到底部（不回到开头）。"""
        self.message_edit.setPlainText(text)
        # 延迟到文本渲染完成后再滚动，确保滚动条 maximum 已更新
        from PySide6.QtCore import QTimer

        QTimer.singleShot(0, self._scroll_chat_to_bottom)

    def _scroll_chat_to_bottom(self) -> None:
        bar = self.message_edit.verticalScrollBar()
        bar.setValue(bar.maximum())

    def _cleanup_old_data(self) -> None:
        """启动时清理过期会话和检索缓存。"""
        async def _cleanup():
            async with async_session_factory() as db:
                count = await ConversationService.cleanup_expired(db)
                await db.commit()
                return count

        try:
            count = asyncio.run(_cleanup())
            if count > 0:
                self.statusBar().showMessage(f"已清理 {count} 个过期对话", 3000)
        except Exception:
            pass

    def _invalidate_retrieval_cache(self) -> None:
        """知识库变更后使检索缓存失效。"""
        async def _invalidate():
            async with async_session_factory() as db:
                await ConversationService.invalidate_retrieval_cache(db)
                await db.commit()

        try:
            asyncio.run(_invalidate())
        except Exception:
            pass

    def _new_conversation(self) -> None:
        async def _create():
            async with async_session_factory() as db:
                conv = await ConversationService.create_conversation(db)
                await db.commit()
                return conv

        try:
            conv = asyncio.run(_create())
            self._current_conv_id = conv.id
            self._load_conversations()
            self._set_chat_text("新对话已创建，开始提问吧。")
            self.query_input.setFocus()
        except Exception as exc:
            self.statusBar().showMessage(f"创建失败: {exc}")

    def _conversation_selected(self, row: int) -> None:
        if row < 0 or row >= len(self._conversations):
            return
        conv = self._conversations[row]
        self._current_conv_id = conv.id

        async def _load_msgs():
            async with async_session_factory() as db:
                return await ConversationService.get_messages(db, conv.id, limit=50)

        try:
            messages = asyncio.run(_load_msgs())
        except Exception:
            messages = []

        lines = []
        for m in messages:
            if m.role == "user" and m.original_query:
                rw = f" [改写: {m.rewritten_query}]" if m.rewritten_query else ""
                lines.append(f"🙂 你: {m.original_query}{rw}")
            elif m.role == "assistant":
                text = m.answer or m.error or "(无回答)"
                lines.append(f"🤖 助手: {text}\n")
        self._set_chat_text("\n\n".join(lines) if lines else "开始新对话")

    def _delete_conversation(self) -> None:
        row = self.conv_list.currentRow()
        if row < 0 or row >= len(self._conversations):
            return
        conv = self._conversations[row]
        reply = QMessageBox.question(
            self,
            "确认删除",
            f'删除会话 "{conv.title}"？',
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        async def _delete():
            async with async_session_factory() as db:
                ok = await ConversationService.delete_conversation(db, conv.id)
                await db.commit()
                return ok

        try:
            asyncio.run(_delete())
            if self._current_conv_id == conv.id:
                self._current_conv_id = None
                self._set_chat_text("选择或创建对话开始提问")
            self._load_conversations()
        except Exception as exc:
            self.statusBar().showMessage(f"删除失败: {exc}")

    def _filter_conversations(self, keyword: str) -> None:
        keyword = keyword.strip()
        if not keyword:
            self._load_conversations()
            return

        async def _search():
            async with async_session_factory() as db:
                return await ConversationService.search_conversations(db, keyword)

        try:
            results = asyncio.run(_search())
        except Exception:
            return

        self.conv_list.blockSignals(True)
        self.conv_list.clear()
        for conv in results:
            self.conv_list.addItem(f"{conv.title}  ({conv.message_count})")
        self.conv_list.blockSignals(False)

    def _export_conversation(self) -> None:
        if not self._current_conv_id:
            QMessageBox.information(self, "未选择", "请先选择一个对话。")
            return

        async def _export():
            async with async_session_factory() as db:
                return await ConversationService.export_conversation(
                    db, self._current_conv_id
                )

        try:
            data = asyncio.run(_export())
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", str(exc))
            return

        if data is None:
            QMessageBox.information(self, "导出失败", "会话不存在。")
            return

        import json
        dest, _ = QFileDialog.getSaveFileName(
            self,
            "导出对话",
            f"{data.get('title', 'conversation')}.json",
            "JSON (*.json)",
        )
        if not dest:
            return
        try:
            Path(dest).write_text(
                json.dumps(data, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            self.statusBar().showMessage(f"已导出到 {dest}", 5000)
        except Exception as exc:
            QMessageBox.critical(self, "保存失败", str(exc))

    def _clear_all_conversations(self) -> None:
        reply = QMessageBox.question(
            self,
            "确认清空",
            "确定要删除所有对话记录？此操作不可撤销。",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        async def _clear():
            async with async_session_factory() as db:
                count = await ConversationService.clear_all_conversations(db)
                await db.commit()
                return count

        try:
            count = asyncio.run(_clear())
            self._current_conv_id = None
            self._conversations = []
            self.conv_list.clear()
            self._set_chat_text("所有对话已清空。")
            self.start_disk_stats()
            self.statusBar().showMessage(f"已清空 {count} 个对话", 5000)
        except Exception as exc:
            self.statusBar().showMessage(f"清空失败: {exc}")

    def _build_knowledge_tab(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(12)

        heading = QLabel("知识库文档")
        heading.setObjectName("pageTitle")
        layout.addWidget(heading)

        actions = QHBoxLayout()
        self.add_files_button = QPushButton("添加文件")
        self.add_files_button.setIcon(self.style().standardIcon(QStyle.SP_DialogOpenButton))
        self.add_files_button.clicked.connect(self.choose_files)
        actions.addWidget(self.add_files_button)

        self.add_folder_button = QPushButton("添加文件夹")
        self.add_folder_button.setIcon(self.style().standardIcon(QStyle.SP_DirOpenIcon))
        self.add_folder_button.clicked.connect(self.choose_folder)
        actions.addWidget(self.add_folder_button)

        self.rebuild_button = QPushButton("从文件夹重建")
        self.rebuild_button.setIcon(
            self.style().standardIcon(QStyle.SP_BrowserReload)
        )
        self.rebuild_button.clicked.connect(self.choose_rebuild_folder)
        actions.addWidget(self.rebuild_button)

        self.diagnose_button = QPushButton("诊断")
        self.diagnose_button.setIcon(
            self.style().standardIcon(QStyle.SP_MessageBoxInformation)
        )
        self.diagnose_button.clicked.connect(self.start_index_diagnosis)
        actions.addWidget(self.diagnose_button)

        self.data_dir_button = QPushButton("数据目录")
        self.data_dir_button.setIcon(
            self.style().standardIcon(QStyle.SP_DirOpenIcon)
        )
        self.data_dir_button.clicked.connect(self.open_data_directory)
        actions.addWidget(self.data_dir_button)

        actions.addStretch()

        self.delete_button = QPushButton("删除所选")
        self.delete_button.setObjectName("dangerButton")
        self.delete_button.setIcon(self.style().standardIcon(QStyle.SP_TrashIcon))
        self.delete_button.clicked.connect(self.delete_selected_document)
        actions.addWidget(self.delete_button)
        layout.addLayout(actions)

        progress_row = QHBoxLayout()
        self.index_progress_label = QLabel("索引任务")
        self.index_progress_label.setObjectName("mutedText")
        self.index_progress_label.setVisible(False)
        progress_row.addWidget(self.index_progress_label)
        self.index_progress = QProgressBar()
        self.index_progress.setRange(0, 100)
        self.index_progress.setValue(0)
        self.index_progress.setTextVisible(True)
        self.index_progress.setVisible(False)
        progress_row.addWidget(self.index_progress, 1)
        self.cancel_index_button = QPushButton("取消")
        self.cancel_index_button.setIcon(
            self.style().standardIcon(QStyle.SP_BrowserStop)
        )
        self.cancel_index_button.setDisabled(True)
        self.cancel_index_button.clicked.connect(self.cancel_indexing)
        progress_row.addWidget(self.cancel_index_button)
        layout.addLayout(progress_row)

        self.documents_table = QTableWidget(0, 4)
        self.documents_table.setHorizontalHeaderLabels(["文件名", "页数/表数", "节点", "片段"])
        self.documents_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.documents_table.setSelectionMode(QTableWidget.SingleSelection)
        self.documents_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.documents_table.verticalHeader().setVisible(False)
        header = self.documents_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        for index in range(1, 4):
            header.setSectionResizeMode(index, QHeaderView.ResizeToContents)
        layout.addWidget(self.documents_table, 1)

        self.index_path_label = QLabel()
        self.index_path_label.setObjectName("mutedText")
        self.index_path_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        layout.addWidget(self.index_path_label)

        disk_row = QHBoxLayout()
        self.disk_usage_label = QLabel("磁盘占用：计算中…")
        self.disk_usage_label.setObjectName("mutedText")
        disk_row.addWidget(self.disk_usage_label)
        disk_row.addStretch()
        self.refresh_disk_button = QPushButton("刷新")
        self.refresh_disk_button.setMaximumWidth(64)
        self.refresh_disk_button.clicked.connect(self.start_disk_stats)
        disk_row.addWidget(self.refresh_disk_button)
        layout.addLayout(disk_row)
        return page

    def _build_settings_tab(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(12)

        heading = QLabel("模型设置")
        heading.setObjectName("pageTitle")
        layout.addWidget(heading)

        llm_heading = QLabel("LLM")
        llm_heading.setObjectName("sectionTitle")
        layout.addWidget(llm_heading)

        llm_form = QFormLayout()
        llm_form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        self.api_key_input = QLineEdit()
        self.api_key_input.setEchoMode(QLineEdit.Password)
        self.api_key_input.setPlaceholderText("填写你的 LLM API Key")
        llm_form.addRow("API Key", self.api_key_input)

        self.base_url_input = QLineEdit()
        self.base_url_input.setPlaceholderText("https://api.deepseek.com")
        llm_form.addRow("Base URL", self.base_url_input)

        self.model_input = QLineEdit()
        self.model_input.setPlaceholderText("deepseek-v4-flash")
        llm_form.addRow("Model", self.model_input)
        layout.addLayout(llm_form)

        retrieval_heading = QLabel("远程检索")
        retrieval_heading.setObjectName("sectionTitle")
        layout.addWidget(retrieval_heading)

        retrieval_form = QFormLayout()
        retrieval_form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        self.retrieval_api_key_input = QLineEdit()
        self.retrieval_api_key_input.setEchoMode(QLineEdit.Password)
        self.retrieval_api_key_input.setPlaceholderText("填写 Embedding / Reranker API Key")
        retrieval_form.addRow("API Key", self.retrieval_api_key_input)

        self.retrieval_base_url_input = QLineEdit()
        self.retrieval_base_url_input.setPlaceholderText(DEFAULT_RETRIEVAL_BASE_URL)
        retrieval_form.addRow("Base URL", self.retrieval_base_url_input)

        self.embedding_model_input = QLineEdit()
        self.embedding_model_input.setPlaceholderText(DEFAULT_EMBEDDING_MODEL)
        retrieval_form.addRow("Embedding Model", self.embedding_model_input)

        self.reranker_model_input = QLineEdit()
        self.reranker_model_input.setPlaceholderText(DEFAULT_RERANKER_MODEL)
        retrieval_form.addRow("Reranker Model", self.reranker_model_input)
        layout.addLayout(retrieval_form)

        actions = QHBoxLayout()
        self.save_feedback = QLabel("已加载保存的设置")
        self.save_feedback.setObjectName("mutedText")
        actions.addWidget(self.save_feedback)
        actions.addStretch()
        self.save_button = QPushButton("设置已保存")
        self.save_button.setIcon(self.style().standardIcon(QStyle.SP_DialogSaveButton))
        self.save_button.setDisabled(True)
        self.save_button.clicked.connect(self.save_settings)
        actions.addWidget(self.save_button)
        layout.addLayout(actions)

        privacy_heading = QLabel("隐私与安全")
        privacy_heading.setObjectName("sectionTitle")
        layout.addWidget(privacy_heading)

        self.offline_checkbox = QCheckBox("完全离线模式（禁止一切远程请求）")
        self.offline_checkbox.setChecked(True)
        layout.addWidget(self.offline_checkbox)

        self.offline_note = QLabel(
            "离线模式下不会发起任何远程请求，检索与回答完全在本地完成。"
        )
        self.offline_note.setWordWrap(True)
        self.offline_note.setObjectName("mutedText")
        layout.addWidget(self.offline_note)

        self.credential_backend_label = QLabel(
            f"API Key 安全存储：{get_backend_name()}"
        )
        self.credential_backend_label.setWordWrap(True)
        self.credential_backend_label.setObjectName("mutedText")
        layout.addWidget(self.credential_backend_label)

        note = QLabel(
            "启用远程检索时，只会发送明确展示的问题和检索到的文档片段到配置的服务地址；"
            "API Key 保存在系统凭据库，不会写入明文设置文件。"
        )
        note.setWordWrap(True)
        note.setObjectName("mutedText")
        layout.addWidget(note)
        layout.addStretch()

        for line_edit in (
            self.api_key_input,
            self.base_url_input,
            self.model_input,
            self.retrieval_api_key_input,
            self.retrieval_base_url_input,
            self.embedding_model_input,
            self.reranker_model_input,
        ):
            line_edit.textChanged.connect(self._mark_settings_dirty)
        for checkbox in (
            self.use_llm_checkbox,
            self.use_embedding_checkbox,
            self.use_reranker_checkbox,
            self.offline_checkbox,
        ):
            checkbox.stateChanged.connect(self._mark_settings_dirty)
        self.offline_checkbox.stateChanged.connect(
            lambda *_args: self._apply_offline_state()
        )
        return page

    def _load_settings(self) -> None:
        self._settings_loading = True
        try:
            self.api_key_input.setText(get_secret(DEFAULT_SERVICE, "llm_api_key"))
            self.base_url_input.setText(
                self.settings.value("llm/base_url", "https://api.deepseek.com", str)
            )
            self.model_input.setText(
                normalize_llm_model(
                    self.settings.value("llm/model", DEFAULT_LLM_MODEL, str)
                )
            )
            self.retrieval_api_key_input.setText(
                get_secret(DEFAULT_SERVICE, "retrieval_api_key")
            )
            self.retrieval_base_url_input.setText(
                self.settings.value(
                    "retrieval/base_url",
                    DEFAULT_RETRIEVAL_BASE_URL,
                    str,
                )
            )
            self.embedding_model_input.setText(
                self.settings.value(
                    "retrieval/embedding_model",
                    DEFAULT_EMBEDDING_MODEL,
                    str,
                )
            )
            self.reranker_model_input.setText(
                self.settings.value(
                    "retrieval/reranker_model",
                    DEFAULT_RERANKER_MODEL,
                    str,
                )
            )
            self.use_llm_checkbox.setChecked(
                self.settings.value("features/use_llm", True, bool)
            )
            self.use_embedding_checkbox.setChecked(
                self.settings.value("features/use_embedding", False, bool)
            )
            self.use_reranker_checkbox.setChecked(
                self.settings.value("features/use_reranker", False, bool)
            )
            self.offline_checkbox.setChecked(
                self.settings.value("privacy/offline_mode", True, bool)
            )
        finally:
            self._settings_loading = False
        self._apply_offline_state()
        self._set_settings_saved_state("已加载保存的设置")

    def save_settings(self) -> None:
        credential_errors = []
        try:
            set_secret(DEFAULT_SERVICE, "llm_api_key", self.api_key_input.text().strip())
            set_secret(
                DEFAULT_SERVICE,
                "retrieval_api_key",
                self.retrieval_api_key_input.text().strip(),
            )
        except Exception as exc:
            credential_errors.append(str(exc))
        self.settings.setValue("llm/base_url", self.base_url_input.text().strip())
        normalized_model = normalize_llm_model(self.model_input.text())
        self.settings.setValue("llm/model", normalized_model)
        self.model_input.setText(normalized_model)
        self.settings.setValue(
            "retrieval/base_url",
            self.retrieval_base_url_input.text().strip(),
        )
        self.settings.setValue(
            "retrieval/embedding_model",
            self.embedding_model_input.text().strip(),
        )
        self.settings.setValue(
            "retrieval/reranker_model",
            self.reranker_model_input.text().strip(),
        )
        self.settings.setValue("features/use_llm", self.use_llm_checkbox.isChecked())
        self.settings.setValue(
            "features/use_embedding",
            self.use_embedding_checkbox.isChecked(),
        )
        self.settings.setValue(
            "features/use_reranker",
            self.use_reranker_checkbox.isChecked(),
        )
        self.settings.setValue("privacy/offline_mode", self.offline_checkbox.isChecked())
        self.settings.sync()
        self._apply_offline_state()
        if credential_errors:
            QMessageBox.warning(
                self,
                "凭据保存失败",
                "API Key 未能写入系统凭据库，请检查系统凭据服务后重试：\n"
                + "\n".join(credential_errors),
            )
        self._set_settings_saved_state("设置已保存到当前系统用户")
        self.statusBar().showMessage("模型设置已保存", 4000)

    def _mark_settings_dirty(self, *_args) -> None:
        if self._settings_loading:
            return
        self.save_button.setText("保存设置")
        self.save_button.setEnabled(True)
        self.save_feedback.setText("有未保存的更改")
        self.save_feedback.setObjectName("warningText")
        self.save_feedback.style().unpolish(self.save_feedback)
        self.save_feedback.style().polish(self.save_feedback)

    def _set_settings_saved_state(self, message: str) -> None:
        self.save_button.setText("设置已保存")
        self.save_button.setDisabled(True)
        self.save_feedback.setText(message)
        self.save_feedback.setObjectName("savedText")
        self.save_feedback.style().unpolish(self.save_feedback)
        self.save_feedback.style().polish(self.save_feedback)

    def _apply_offline_state(self) -> None:
        offline = self.offline_checkbox.isChecked()
        set_remote_access(not offline)
        for checkbox in (
            self.use_llm_checkbox,
            self.use_embedding_checkbox,
            self.use_reranker_checkbox,
        ):
            checkbox.setEnabled(not offline)
        self.offline_note.setVisible(offline)
        if offline:
            self._set_retrieval_status("模式：— · 完全离线")
            self._network_active = False

    def _set_network_indicator(self, active: bool) -> None:
        self._network_active = bool(active)
        if active:
            self._set_retrieval_status("模式：— · 调用中…")

    def _set_retrieval_status(self, text: str) -> None:
        self.network_status_label.setText(text)

    def _build_retrieval_status(self, retrieval: Dict[str, Any], mode: str) -> str:
        retrieval_mode = str(retrieval.get("mode") or "")
        parts = [f"模式：{self._retrieval_mode_label(retrieval_mode)}"]
        if retrieval.get("offline"):
            parts.append("完全离线")
        elif mode in ("llm_error", "embedding_error", "reranker_error"):
            parts.append("失败")
        elif retrieval.get("cache_hit"):
            parts.append("缓存命中")
        elif retrieval.get("remote"):
            parts.append("远程调用")
        else:
            parts.append("本地")
        return " · ".join(parts)

    @staticmethod
    def _retrieval_mode_label(mode: str) -> str:
        return {
            "bm25": "BM25",
            "hybrid": "混合检索",
            "bm25_rerank": "BM25 + Reranker",
            "hybrid_rerank": "混合 + Reranker",
            "structured": "结构化计算",
            "summary": "文档概览",
            "inventory": "文件清单",
            "mixed": "混合计算",
        }.get(mode, mode or "—")

    def _consented_endpoints(self) -> set[str]:
        raw = str(self.settings.value("privacy/remote_consent", "", str) or "")
        return {value for value in raw.split(",") if value}

    def _confirm_remote_consent(
        self,
        query: str,
        use_llm: bool,
        use_embedding: bool,
        use_reranker: bool,
    ) -> bool:
        endpoints: list[tuple[str, str]] = []
        llm_base = self.base_url_input.text().strip()
        if use_llm and llm_base:
            endpoints.append(("LLM", llm_base))
        retrieval_base = self.retrieval_base_url_input.text().strip()
        if (use_embedding or use_reranker) and retrieval_base:
            endpoints.append(("Embedding / Reranker", retrieval_base))
        if not endpoints:
            return True

        consented = self._consented_endpoints()
        pending = [(name, url) for name, url in endpoints if url not in consented]
        if not pending:
            return True

        endpoint_text = "\n".join(f"• {name}：{url}" for name, url in pending)
        message = (
            "即将发起远程调用，将把以下数据发送到配置的服务：\n\n"
            f"问题内容：{query[:200]}\n\n"
            "发送范围：问题文本，以及本地检索到的文档片段、文件名和 "
            "Sheet/页面定位元数据。\n\n"
            "目标服务：\n"
            f"{endpoint_text}\n\n"
            "请确认允许向上述服务发送这些数据。"
        )
        box = QMessageBox(self)
        box.setWindowTitle("确认远程调用")
        box.setIcon(QMessageBox.Warning)
        box.setText(message)
        remember = box.addButton("同意并记住", QMessageBox.AcceptRole)
        once = box.addButton("仅本次同意", QMessageBox.AcceptRole)
        cancel = box.addButton("取消", QMessageBox.RejectRole)
        box.setDefaultButton(cancel)
        box.exec()
        clicked = box.clickedButton()
        if clicked is cancel:
            return False
        if clicked is remember:
            merged = consented | {url for _, url in pending}
            self.settings.setValue(
                "privacy/remote_consent",
                ",".join(sorted(merged)),
            )
            self.settings.sync()
        return True

    def choose_files(self) -> None:
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "选择知识文档",
            "",
            DESKTOP_FILE_FILTER,
        )
        if paths:
            self.start_indexing([Path(path) for path in paths])

    def choose_folder(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "选择知识文档文件夹")
        if not folder:
            return
        paths = [
            path
            for path in sorted(Path(folder).rglob("*"))
            if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
        ]
        if not paths:
            QMessageBox.information(
                self,
                "没有文档",
                "所选文件夹中没有 PDF、TXT、MD、CSV 或 XLSX 文件。",
            )
            return
        self.start_indexing(paths, source_root=Path(folder))

    def choose_rebuild_folder(self) -> None:
        folder = QFileDialog.getExistingDirectory(
            self,
            "选择完整知识库文件夹",
        )
        if not folder:
            return
        root = Path(folder)
        paths = [
            path
            for path in sorted(root.rglob("*"))
            if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
        ]
        if not paths:
            QMessageBox.information(
                self,
                "没有文档",
                "所选文件夹中没有可重建的知识文档。",
            )
            return
        answer = QMessageBox.question(
            self,
            "确认重建",
            "重建成功后，当前知识库将完全替换为所选文件夹中的文档。继续？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if answer == QMessageBox.Yes:
            self.start_indexing(
                paths,
                source_root=root,
                force_reparse=True,
                replace_all=True,
            )

    def start_indexing(
        self,
        paths: List[Path],
        csv_encodings: Dict[str, str] | None = None,
        *,
        source_root: Path | None = None,
        force_reparse: bool = False,
        replace_all: bool = False,
    ) -> None:
        self._set_busy(True, "正在读取并构建索引...")
        self._set_indexing_state(True)
        resolved_encodings = dict(csv_encodings or {})
        worker = IndexWorker(
            paths,
            self.index_dir,
            resolved_encodings,
            source_root=source_root,
            force_reparse=force_reparse,
            replace_all=replace_all,
        )
        self._active_index_worker = worker
        worker.completed.connect(self._indexing_completed)
        worker.progress_changed.connect(self._indexing_progress_changed)
        worker.encoding_required.connect(
            lambda failed_path, message: self._choose_csv_encoding(
                paths,
                resolved_encodings,
                failed_path,
                message,
                source_root=source_root,
                force_reparse=force_reparse,
                replace_all=replace_all,
            )
        )
        worker.cancelled.connect(self._indexing_cancelled)
        worker.failed.connect(self._indexing_failed)
        worker.finished.connect(lambda: self._release_worker(worker))
        self._workers.append(worker)
        worker.start()

    def _choose_csv_encoding(
        self,
        paths: List[Path],
        csv_encodings: Dict[str, str],
        failed_path: str,
        message: str,
        *,
        source_root: Path | None = None,
        force_reparse: bool = False,
        replace_all: bool = False,
    ) -> None:
        self._set_busy(False)
        self._set_indexing_state(False)
        choices = list(CSV_ENCODING_CANDIDATES) + ["latin-1"]
        encoding, accepted = QInputDialog.getItem(
            self,
            "选择 CSV 编码",
            f"{Path(failed_path).name} 无法自动确认编码。\n{message}\n请选择编码后重试：",
            choices,
            0,
            True,
        )
        if not accepted or not encoding.strip():
            self.statusBar().showMessage("已取消 CSV 编码选择", 7000)
            return
        retry_encodings = dict(csv_encodings)
        retry_encodings[str(Path(failed_path).resolve())] = encoding.strip()
        self.start_indexing(
            paths,
            retry_encodings,
            source_root=source_root,
            force_reparse=force_reparse,
            replace_all=replace_all,
        )

    def _indexing_completed(self, payload: Dict[str, Any]) -> None:
        self._set_indexing_state(False)
        self._set_busy(False)
        self.refresh_documents()
        self._invalidate_retrieval_cache()
        self.start_disk_stats()
        added = int(payload.get("added_count") or 0)
        updated = int(payload.get("updated_count") or 0)
        removed = int(payload.get("removed_count") or 0)
        skipped = payload.get("skipped_files") or []
        failed = payload.get("failed_files") or []
        parts = [f"新增 {added}", f"更新 {updated}", f"删除 {removed}"]
        if skipped:
            parts.append(f"未变化 {len(skipped)}")
        if failed:
            parts.append(f"失败 {len(failed)}")
            details = "\n".join(
                f"{item.get('filename')}: {item.get('error')}"
                for item in failed
            )
            QMessageBox.warning(
                self,
                "部分文档未更新",
                "其他文档已提交，以下文档保留旧版本或未加入：\n" + details,
            )
        message = "，".join(parts)
        self.statusBar().showMessage(message, 7000)

    def _indexing_progress_changed(self, payload: Dict[str, Any]) -> None:
        current = int(payload.get("current") or 0)
        total = max(int(payload.get("total") or 0), 1)
        phase = str(payload.get("phase") or "")
        filename = str(payload.get("filename") or "")
        labels = {
            "fingerprint": "检查",
            "parsing": "解析",
            "parsed": "已解析",
            "skipped": "未变化",
            "failed": "失败",
            "committing": "提交",
            "completed": "完成",
        }
        self.index_progress.setValue(min(int(current * 100 / total), 100))
        text = labels.get(phase, "索引")
        self.index_progress_label.setText(
            f"{text}：{filename}" if filename else text
        )

    def cancel_indexing(self) -> None:
        worker = self._active_index_worker
        if worker is None or not worker.isRunning():
            return
        worker.requestInterruption()
        self.cancel_index_button.setDisabled(True)
        self.index_progress_label.setText("正在取消...")
        self.statusBar().showMessage("正在取消索引任务")

    def _indexing_cancelled(self) -> None:
        self._set_indexing_state(False)
        self._set_busy(False)
        self.statusBar().showMessage("索引任务已取消，原索引保持不变", 7000)

    def _indexing_failed(self, message: str) -> None:
        self._set_indexing_state(False)
        self._set_busy(False)
        QMessageBox.critical(self, "索引失败", message)
        self.statusBar().showMessage("索引失败，原索引保持不变", 7000)

    def start_index_diagnosis(self) -> None:
        self._set_busy(True, "正在诊断索引...")
        worker = DiagnosticWorker(self.index_dir)
        worker.completed.connect(self._diagnosis_completed)
        worker.failed.connect(self._task_failed)
        worker.finished.connect(lambda: self._release_worker(worker))
        self._workers.append(worker)
        worker.start()

    def _diagnosis_completed(self, payload: Dict[str, Any]) -> None:
        self._set_busy(False)
        counts = payload.get("counts") or {}
        issues = payload.get("issues") or []
        warnings = payload.get("warnings") or []
        lines = [
            f"状态：{payload.get('status')}",
            (
                "文档 {documents}，节点 {nodes}，父节点 {parents}，Chunk {chunks}"
            ).format(**counts),
        ]
        if payload.get("recovered_transaction"):
            lines.append("已自动恢复上次未完成的索引事务。")
        lines.extend(f"错误：{item.get('message')}" for item in issues)
        lines.extend(f"警告：{item.get('message')}" for item in warnings)
        QMessageBox.information(self, "索引诊断", "\n".join(lines))

    def _data_root(self) -> Path:
        """应用数据根目录：EXE 用 AppDataLocation，开发用仓库 data。"""
        if getattr(sys, "frozen", False):
            return Path(
                QStandardPaths.writableLocation(QStandardPaths.AppDataLocation)
            ).resolve()
        return Path("data").resolve()

    def open_data_directory(self) -> None:
        root = self._data_root()
        try:
            root.mkdir(parents=True, exist_ok=True)
        except OSError:
            pass
        QDesktopServices.openUrl(QUrl.fromLocalFile(str(root)))

    def start_disk_stats(self) -> None:
        """后台统计索引与聊天数据库磁盘占用。"""
        self.disk_usage_label.setText("磁盘占用：计算中…")
        worker = DiskStatsWorker(self.index_dir, _desktop_db_path())
        worker.completed.connect(self._disk_stats_completed)
        worker.finished.connect(lambda: self._release_worker(worker))
        self._workers.append(worker)
        worker.start()

    def _disk_stats_completed(self, payload: Dict[str, Any]) -> None:
        index_bytes = int(payload.get("index_bytes") or 0)
        db_bytes = int(payload.get("db_bytes") or 0)
        self.disk_usage_label.setText(
            f"磁盘占用：索引 {_human_size(index_bytes)} · "
            f"聊天数据库 {_human_size(db_bytes)}"
        )

    def _structure_label(self, document: dict) -> str:
        """PDF 页数 / Excel 表数；旧索引缺精确统计时显示"未知"需重建。"""
        page_count = document.get("page_count")
        sheet_count = document.get("sheet_count")
        if page_count is not None:
            return f"{page_count} 页"
        if sheet_count is not None:
            return f"{sheet_count} 表"
        filename = str(document.get("filename") or "").casefold()
        if filename.endswith((".pdf", ".xlsx", ".xls", ".csv")):
            return "未知"
        return "—"

    def refresh_documents(self) -> None:
        index_error = ""
        try:
            documents = list_index_documents(self.index_dir)
        except IndexFormatError as exc:
            documents = []
            index_error = str(exc)
        self.documents_table.setRowCount(len(documents))
        for row, document in enumerate(documents):
            filename_item = QTableWidgetItem(str(document.get("filename") or ""))
            filename_item.setData(Qt.UserRole, str(document.get("filename") or ""))
            self.documents_table.setItem(row, 0, filename_item)
            self.documents_table.setItem(
                row, 1, QTableWidgetItem(self._structure_label(document))
            )
            self.documents_table.setItem(
                row, 2, QTableWidgetItem(str(document.get("node_count") or 0))
            )
            self.documents_table.setItem(
                row, 3, QTableWidgetItem(str(document.get("chunk_count") or 0))
            )
        label = f"索引位置：{self.index_dir}"
        if index_error:
            label += f"\n索引需要重建：{index_error}"
            self.statusBar().showMessage("旧索引不兼容，请重建索引")
        self.index_path_label.setText(label)
        self.ask_button.setEnabled(bool(documents))
        self.delete_button.setEnabled(bool(documents))

    def delete_selected_document(self) -> None:
        row = self.documents_table.currentRow()
        if row < 0:
            QMessageBox.information(self, "请选择文档", "请先选择要删除的文档。")
            return
        item = self.documents_table.item(row, 0)
        filename = str(item.data(Qt.UserRole) or item.text())
        answer = QMessageBox.question(
            self,
            "删除文档",
            f"删除“{filename}”及其索引？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            return
        try:
            delete_index_document(filename, self.index_dir)
            self.refresh_documents()
            self._invalidate_retrieval_cache()
            self.start_disk_stats()
            self.clear_sources()
            self._set_chat_text("文档已删除。")
            self.statusBar().showMessage("文档及索引已删除", 5000)
        except Exception as exc:
            QMessageBox.critical(self, "删除失败", str(exc))

    def ask_question(self) -> None:
        query = self.query_input.toPlainText().strip()
        if not query:
            QMessageBox.information(self, "请输入问题", "请先输入要查询的问题。")
            return

        offline = self.offline_checkbox.isChecked()
        use_llm = self.use_llm_checkbox.isChecked()
        use_embedding = self.use_embedding_checkbox.isChecked()
        use_reranker = self.use_reranker_checkbox.isChecked()

        # 远程需求判断基于原问题（改写已移入后台线程，不阻塞 UI）。
        try:
            planning_documents = list_index_documents(self.index_dir)
        except IndexFormatError:
            planning_documents = []
        query_plan = plan_query(query, planning_documents)
        if query_plan.is_structured_inventory:
            remote_use_llm = False
            remote_use_embedding = False
            remote_use_reranker = False
        elif query_plan.is_summary:
            remote_use_llm = use_llm
            remote_use_embedding = False
            remote_use_reranker = False
        else:
            remote_use_llm = use_llm
            remote_use_embedding = use_embedding
            remote_use_reranker = use_reranker
        remote_requested = (
            remote_use_llm
            or remote_use_embedding
            or remote_use_reranker
        ) and not offline
        if remote_requested and not self._confirm_remote_consent(
            query,
            remote_use_llm,
            remote_use_embedding,
            remote_use_reranker,
        ):
            return

        self.save_settings()
        self._set_busy(True, "正在检索并生成答案...")
        self.clear_sources()
        self._set_network_indicator(remote_requested)

        # Show user question in chat
        current = self.message_edit.toPlainText()
        if current:
            current += "\n\n"
        self._set_chat_text(current + f"🙂 你: {query}\n\n🤖 助手: 正在查询...")

        worker = QueryWorker(
            query,
            self.index_dir,
            conv_id=self._current_conv_id,
            use_llm=use_llm,
            llm_api_key=self.api_key_input.text().strip(),
            llm_base_url=self.base_url_input.text().strip(),
            llm_model=normalize_llm_model(self.model_input.text()),
            use_embedding=use_embedding,
            use_reranker=use_reranker,
            retrieval_api_key=self.retrieval_api_key_input.text().strip(),
            retrieval_base_url=self.retrieval_base_url_input.text().strip(),
            embedding_model=self.embedding_model_input.text().strip(),
            reranker_model=self.reranker_model_input.text().strip(),
            offline=offline,
        )
        worker.completed.connect(self._query_completed)
        worker.failed.connect(self._task_failed)
        worker.finished.connect(lambda: self._release_worker(worker))
        self._workers.append(worker)
        worker.start()
        self.query_input.clear()

    def _query_completed(self, payload: Dict[str, Any]) -> None:
        self._set_busy(False)
        self._set_network_indicator(False)
        answer = str(payload.get("answer") or "没有返回答案。")
        sources = payload.get("sources") or []
        mode = str(payload.get("mode") or "")
        retrieval = payload.get("retrieval") or {}

        # 更新实际检索模式与网络状态（持久状态行）
        self._set_retrieval_status(self._build_retrieval_status(retrieval, mode))

        # 后台 worker 已创建/复用会话并保存消息，这里只同步 UI 状态
        conv_id = payload.get("conv_id") or self._current_conv_id
        if conv_id:
            self._current_conv_id = conv_id

        # Show sources
        self.render_sources(sources)

        # Replace "正在查询..." in chat
        current = self.message_edit.toPlainText()
        current = current.replace("正在查询...", answer)
        self._set_chat_text(current)

        # Reload conversation list (updates message count)
        self._load_conversations()

        if retrieval.get("offline"):
            self.statusBar().showMessage("完全离线模式，本地回答完成", 7000)
        elif mode == "llm_error":
            self.statusBar().showMessage("LLM 配置或请求失败", 7000)
        elif mode == "embedding_error":
            self.statusBar().showMessage("Embedding 配置或请求失败", 7000)
        elif mode == "reranker_error":
            self.statusBar().showMessage("Reranker 配置或请求失败", 7000)
        elif mode == "llm":
            self.statusBar().showMessage("LLM 回答完成", 5000)
        elif retrieval.get("remote"):
            self.statusBar().showMessage("远程检索完成", 5000)
        else:
            self.statusBar().showMessage("本地检索完成", 5000)

    def clear_sources(self) -> None:
        while self.sources_layout.count() > 1:
            item = self.sources_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

    def render_sources(self, sources: List[Dict[str, Any]]) -> None:
        self.clear_sources()
        if not sources:
            empty = QLabel("没有引用来源。")
            empty.setObjectName("mutedText")
            self.sources_layout.insertWidget(0, empty)
            return

        for index, source in enumerate(sources):
            card = QFrame()
            card.setObjectName("sourceBox")
            card_layout = QVBoxLayout(card)
            card_layout.setContentsMargins(10, 8, 10, 8)
            card_layout.setSpacing(6)

            header = QHBoxLayout()
            location = self._source_location_text(source)
            title_text = str(source.get("filename") or "未知文件")
            if location:
                title_text += f"  ·  {location}"
            title = QLabel(title_text)
            title.setObjectName("sourceFilename")
            title.setTextInteractionFlags(Qt.TextSelectableByMouse)
            header.addWidget(title, 1)

            copy_btn = QPushButton("复制")
            copy_btn.setMaximumWidth(52)
            copy_btn.clicked.connect(
                lambda _=False, s=source: self._copy_source(s)
            )
            header.addWidget(copy_btn)

            locate_btn = QPushButton("定位")
            locate_btn.setMaximumWidth(52)
            locate_btn.clicked.connect(
                lambda _=False, s=source: self._locate_source(s)
            )
            header.addWidget(locate_btn)

            toggle_btn = QPushButton("收起" if index == 0 else "展开")
            toggle_btn.setMaximumWidth(52)
            header.addWidget(toggle_btn)
            card_layout.addLayout(header)

            content = QPlainTextEdit()
            content.setReadOnly(True)
            content.setPlainText(
                str(source.get("display_content") or source.get("content") or "")
            )
            content.setMaximumHeight(125)
            content.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            card_layout.addWidget(content)

            def _make_toggle(button: QPushButton, editor: QPlainTextEdit) -> Callable:
                def _toggle() -> None:
                    visible = editor.isVisible()
                    editor.setVisible(not visible)
                    button.setText("收起" if not visible else "展开")

                return _toggle

            toggle_btn.clicked.connect(_make_toggle(toggle_btn, content))
            # 第一条默认展开，其余收起
            if index != 0:
                content.setVisible(False)

            self.sources_layout.insertWidget(self.sources_layout.count() - 1, card)

    def _source_location_text(self, source: Dict[str, Any]) -> str:
        anchor = source.get("source_anchor") or {}
        parts = []
        page = anchor.get("page") or source.get("page_or_sheet")
        if page is not None:
            parts.append(f"第 {page} 页")
        sheet = anchor.get("sheet")
        if sheet:
            parts.append(f"Sheet {sheet}")
        row_numbers = anchor.get("row_numbers")
        if row_numbers:
            first = str(row_numbers[0])
            if len(row_numbers) > 1:
                parts.append(f"行 {first}-{row_numbers[-1]}")
            else:
                parts.append(f"行 {first}")
        return " · ".join(parts)

    def _copy_source(self, source: Dict[str, Any]) -> None:
        filename = str(source.get("filename") or "未知文件")
        location = self._source_location_text(source)
        content = str(
            source.get("display_content") or source.get("content") or ""
        )
        markdown = f"**{filename}**"
        if location:
            markdown += f"（{location}）"
        if content:
            markdown += f"\n\n{content}"
        QApplication.clipboard().setText(markdown)
        self.statusBar().showMessage(f"已复制引用：{filename}", 3000)

    def _resolve_origin_path(self, filename: str) -> Path:
        try:
            documents = list_index_documents(self.index_dir)
        except Exception:
            documents = []
        for document in documents:
            if str(document.get("filename") or "") == filename:
                origin = document.get("origin_path")
                if origin:
                    return Path(str(origin))
                break
        return Path(filename)

    def _locate_source(self, source: Dict[str, Any]) -> None:
        filename = str(source.get("filename") or "")
        origin_path = self._resolve_origin_path(filename)
        dialog = SourcePreviewDialog(source, origin_path, self)
        dialog.exec()

    def _task_failed(self, message: str) -> None:
        self._set_busy(False)
        self._set_network_indicator(False)
        # 后台 worker 已负责保存错误消息，这里只更新 UI。
        current = self.message_edit.toPlainText()
        current = current.replace("正在查询...", f"操作失败：{message}")
        self._set_chat_text(current + "\n")
        self._load_conversations()
        self.clear_sources()
        self.statusBar().showMessage("操作失败", 7000)

    def _set_busy(self, busy: bool, message: str = "") -> None:
        self.ask_button.setDisabled(busy)
        self.add_files_button.setDisabled(busy)
        self.add_folder_button.setDisabled(busy)
        self.rebuild_button.setDisabled(busy)
        self.diagnose_button.setDisabled(busy)
        self.delete_button.setDisabled(busy)
        self.conv_new_btn.setDisabled(busy)
        self.conv_delete_btn.setDisabled(busy)
        self.conv_export_btn.setDisabled(busy)
        self.conv_clear_all_btn.setDisabled(busy)
        if message:
            self.statusBar().showMessage(message)
        if not busy:
            self.refresh_documents()

    def _set_indexing_state(self, active: bool) -> None:
        self.index_progress_label.setVisible(active)
        self.index_progress.setVisible(active)
        self.cancel_index_button.setEnabled(active)
        if active:
            self.index_progress.setValue(0)
            self.index_progress_label.setText("准备索引...")
        else:
            self.cancel_index_button.setDisabled(True)

    def _release_worker(self, worker: QThread) -> None:
        if worker is self._active_index_worker:
            self._active_index_worker = None
        if worker in self._workers:
            self._workers.remove(worker)
        worker.deleteLater()

    def closeEvent(self, event: QCloseEvent) -> None:
        active_workers = [worker for worker in self._workers if worker.isRunning()]
        if active_workers:
            answer = QMessageBox.question(
                self,
                "任务仍在运行",
                "当前任务尚未完成，确定退出？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if answer != QMessageBox.Yes:
                event.ignore()
                return
        event.accept()
        cleanup_temp_files()


APP_STYLE = """
QMainWindow, QWidget {
    background: #f5f7fa;
    color: #18202a;
    font-family: "Segoe UI", "Microsoft YaHei", sans-serif;
    font-size: 13px;
}
QTabWidget::pane {
    border: 1px solid #d9e0ea;
    background: #ffffff;
}
QTabBar::tab {
    background: #e9eef5;
    border: 1px solid #d9e0ea;
    border-bottom: none;
    padding: 9px 18px;
}
QTabBar::tab:selected {
    background: #ffffff;
    color: #1d4ed8;
}
QLabel#pageTitle {
    font-size: 20px;
    font-weight: 700;
}
QLabel#sectionTitle {
    font-size: 14px;
    font-weight: 600;
}
QLabel#mutedText {
    color: #687383;
}
QLabel#savedText {
    color: #067647;
}
QLabel#warningText {
    color: #a15c00;
}
QLabel#sourceFilename {
    color: #556274;
    font-size: 12px;
    font-weight: 600;
}
QLabel#networkStatus {
    color: #b45309;
    font-weight: 600;
}
QLineEdit, QPlainTextEdit, QTableWidget {
    background: #ffffff;
    border: 1px solid #cfd8e5;
    border-radius: 5px;
    selection-background-color: #bfdbfe;
    padding: 7px;
}
QListWidget {
    background: #ffffff;
    border: 1px solid #d9e0ea;
    border-radius: 5px;
    outline: none;
}
QListWidget::item {
    padding: 6px 8px;
    border-bottom: 1px solid #f0f2f5;
}
QListWidget::item:selected {
    background: #dbeafe;
    color: #1d4ed8;
}
QListWidget::item:hover {
    background: #f0f4ff;
}
QPushButton {
    background: #2563eb;
    color: #ffffff;
    border: none;
    border-radius: 5px;
    min-height: 32px;
    padding: 0 14px;
    font-weight: 600;
}
QPushButton:hover {
    background: #1d4ed8;
}
QPushButton:disabled {
    background: #aab5c4;
}
QPushButton#dangerButton {
    background: #b42318;
}
QPushButton#dangerButton:hover {
    background: #912018;
}
QFrame#sourceBox {
    background: #ffffff;
    border: 1px solid #d9e0ea;
    border-radius: 6px;
}
QStatusBar {
    background: #ffffff;
    border-top: 1px solid #d9e0ea;
}
QSplitter::handle {
    background: #e9eef5;
    width: 1px;
}
"""


def _write_ui_test_result(results: list[tuple[str, bool]], detail: str = "") -> None:
    """把 UI 测试结果写到 exe 同级目录，供自动化验证。"""
    import json

    path = Path("ui_test_result.json")
    path.write_text(
        json.dumps(
            {"results": results, "detail": detail},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def run_desktop(smoke_test: bool = False, ui_test: bool = False) -> int:
    app = QApplication.instance() or QApplication(sys.argv)
    app.setApplicationName(APP_NAME)
    app.setOrganizationName(ORGANIZATION)
    window = MainWindow()
    window.show()
    if ui_test:
        from PySide6.QtCore import QTimer

        QTimer.singleShot(500, lambda: _run_ui_test(app, window))
    elif smoke_test:
        from PySide6.QtCore import QTimer

        QTimer.singleShot(500, app.quit)
    return app.exec()


def _run_ui_test(app: QApplication, window: "MainWindow") -> None:
    """驱动真实 EXE 的按钮点击，验证聊天功能并写结果文件。"""
    import time

    from PySide6.QtCore import QTimer

    results: list[tuple[str, bool]] = []
    detail_lines: list[str] = []

    def _finish() -> None:
        _write_ui_test_result(results, "\n".join(detail_lines))
        app.quit()

    # 1. 新建会话
    before = window.conv_list.count()
    window.conv_new_btn.click()
    time.sleep(0.5)
    conv_ok = window.conv_list.count() == before + 1 and window._current_conv_id
    results.append(("新建会话", conv_ok))
    detail_lines.append(f"会话列表 {before} -> {window.conv_list.count()}, id={window._current_conv_id}")

    # 2. 若知识库为空，先添加一个测试文档并等待索引完成
    if window.documents_table.rowCount() == 0:
        import tempfile
        from pathlib import Path as _P

        test_doc = _P(tempfile.gettempdir()) / "ekg_ui_test_doc.txt"
        test_doc.write_text(
            "员工请假制度：年假每年 10 天，病假需提供医院证明。报销制度：差旅费需提供发票，"
            "住宿标准每晚不超过 500 元。",
            encoding="utf-8",
        )
        window.start_indexing([test_doc])
        deadline = time.time() + 60
        indexed = False
        while time.time() < deadline:
            app.processEvents()
            if window.documents_table.rowCount() >= 1:
                indexed = True
                break
            time.sleep(0.3)
        results.append(("文档索引", indexed))
        detail_lines.append(f"文档数={window.documents_table.rowCount()}")
        if not indexed:
            detail_lines.append("索引超时，跳过查询")
            _finish()
            return

    # 3. 查询（离线本地回答）
    window.query_input.setPlainText("年假有几天？")
    window.ask_button.click()
    deadline = time.time() + 30
    answer_text = ""
    while time.time() < deadline:
        txt = window.message_edit.toPlainText()
        if "正在查询" not in txt and "助手" in txt:
            answer_text = txt
            break
        app.processEvents()
        time.sleep(0.3)
    results.append(("查询响应", bool(answer_text)))
    results.append(("回答非占位", "正在查询" not in answer_text and "当前知识库没有" not in answer_text))
    results.append(("无模型token显示", "tokens" not in answer_text and "--- " not in answer_text))
    detail_lines.append(f"回答长度={len(answer_text)}")

    # 2b. 再追加一轮对话累积文本，验证滚动到底部
    window.query_input.setPlainText("报销需要什么条件？")
    window.ask_button.click()
    deadline = time.time() + 30
    while time.time() < deadline:
        txt = window.message_edit.toPlainText()
        if "正在查询" not in txt and txt.endswith("助手"):
            break
        if "正在查询" not in txt and "助手" in txt:
            break
        app.processEvents()
        time.sleep(0.3)
    bar = window.message_edit.verticalScrollBar()
    scroll_ok = bar.maximum() > 0 and bar.value() == bar.maximum()
    results.append(("聊天框滚动到底部", scroll_ok))
    detail_lines.append(f"滚动 value={bar.value()} max={bar.maximum()}")

    # 3. DB 持久化验证
    from app.core.database import async_session_factory
    from app.services.conversation_service import ConversationService

    async def _db_check():
        async with async_session_factory() as db:
            msgs = await ConversationService.get_messages(db, window._current_conv_id, limit=10)
            return len(msgs)
    try:
        import asyncio
        n = asyncio.run(_db_check())
        results.append(("消息已持久化", n >= 2))
        detail_lines.append(f"DB 消息数={n}")
    except Exception as exc:
        results.append(("消息已持久化", False))
        detail_lines.append(f"DB 读取异常: {exc}")

    # 4. 按钮状态（都应启用，证明未被禁用）
    btn_ok = all([
        window.ask_button.isEnabled(),
        window.conv_new_btn.isEnabled(),
        window.conv_delete_btn.isEnabled(),
        window.conv_export_btn.isEnabled(),
        window.conv_clear_all_btn.isEnabled(),
    ])
    results.append(("按钮全部启用", btn_ok))
    detail_lines.append(
        f"ask={window.ask_button.isEnabled()} new={window.conv_new_btn.isEnabled()} "
        f"del={window.conv_delete_btn.isEnabled()} export={window.conv_export_btn.isEnabled()} "
        f"clear={window.conv_clear_all_btn.isEnabled()}"
    )

    QTimer.singleShot(100, _finish)
