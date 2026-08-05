from __future__ import annotations

from abc import ABC, abstractmethod
import codecs
import csv
from datetime import date, datetime, time
from io import BytesIO, StringIO
from itertools import zip_longest
from pathlib import Path, PurePosixPath
import re
from typing import Any, BinaryIO, Iterable, Iterator, Optional, TextIO
from xml.etree.ElementTree import iterparse
from zipfile import BadZipFile, ZipFile

from app.documents.node import (
    DocumentNode,
    NodeType,
    document_id_from_source,
)
from app.documents.pdf_parser import (
    PDF_PARSER_VERSION,
    PdfTextLayerError,
    iter_pdf_document_nodes,
)


DOCUMENT_PARSER_INTERFACE_VERSION = "document_parser_v1"
TEXT_PARSER_VERSION = "text_parser_v1"
CSV_PARSER_VERSION = "csv_parser_v2"
XLSX_PARSER_VERSION = "xlsx_parser_v3"
SUPPORTED_DOCUMENT_EXTENSIONS = {".md", ".txt", ".pdf", ".csv", ".xlsx"}
CSV_ROW_GROUP_SIZE = 50
XLSX_ROW_GROUP_SIZE = 50
TEXT_NODE_MAX_CHARS = 4000
CSV_ENCODING_CANDIDATES = ("utf-8-sig", "utf-8", "gb18030")


class ParserError(RuntimeError):
    """Raised when a document cannot be converted to DocumentNode records."""


class ParserDependencyError(ParserError):
    """Raised when an optional format parser dependency is unavailable."""


class CsvEncodingError(ParserError):
    def __init__(
        self,
        source_path: str,
        *,
        attempted: Iterable[str],
        selected: Optional[str] = None,
    ) -> None:
        self.source_path = source_path
        self.attempted = tuple(attempted)
        self.selected = selected
        if selected:
            detail = f"selected encoding {selected!r} could not decode the file"
        else:
            detail = (
                "encoding could not be confirmed from "
                + ", ".join(repr(item) for item in self.attempted)
            )
        super().__init__(
            f"CSV encoding error for {source_path}: {detail}. "
            "Choose the file encoding and retry."
        )


class CsvStructureError(ParserError):
    def __init__(
        self,
        source_path: str,
        *,
        row_number: int,
        expected_columns: int,
        actual_columns: int,
    ) -> None:
        self.source_path = source_path
        self.row_number = row_number
        self.expected_columns = expected_columns
        self.actual_columns = actual_columns
        super().__init__(
            f"CSV structure error for {source_path} at row {row_number}: "
            f"expected {expected_columns} columns, found {actual_columns}."
        )


class DocumentParser(ABC):
    extensions: tuple[str, ...] = ()

    @abstractmethod
    def iter_nodes(
        self,
        path: Path,
        *,
        document_id: str,
        source_path: str,
    ) -> Iterator[DocumentNode]:
        raise NotImplementedError

    @abstractmethod
    def iter_nodes_from_bytes(
        self,
        filename: str,
        content: bytes,
        *,
        document_id: str,
        source_path: str,
    ) -> Iterator[DocumentNode]:
        raise NotImplementedError


class TextParser(DocumentParser):
    extensions = (".md", ".txt")

    def iter_nodes(
        self,
        path: Path,
        *,
        document_id: str,
        source_path: str,
    ) -> Iterator[DocumentNode]:
        with path.open("r", encoding="utf-8", errors="ignore") as reader:
            yield from _iter_text_nodes(
                reader,
                document_id=document_id,
                source_path=source_path,
                file_type=path.suffix.lower(),
            )

    def iter_nodes_from_bytes(
        self,
        filename: str,
        content: bytes,
        *,
        document_id: str,
        source_path: str,
    ) -> Iterator[DocumentNode]:
        text = content.decode("utf-8-sig", errors="ignore")
        yield from _iter_text_nodes(
            StringIO(text),
            document_id=document_id,
            source_path=source_path,
            file_type=PurePosixPath(filename).suffix.lower(),
        )


class PdfParser(DocumentParser):
    extensions = (".pdf",)

    def iter_nodes(
        self,
        path: Path,
        *,
        document_id: str,
        source_path: str,
    ) -> Iterator[DocumentNode]:
        fitz = _load_fitz()
        with fitz.open(path) as document:
            if document.needs_pass:
                raise ParserError(
                    f"PDF is encrypted and requires a password: {source_path}"
                )
            try:
                yield from iter_pdf_document_nodes(
                    document,
                    document_id=document_id,
                    source_path=source_path,
                )
            except PdfTextLayerError as exc:
                raise ParserError(str(exc)) from exc

    def iter_nodes_from_bytes(
        self,
        filename: str,
        content: bytes,
        *,
        document_id: str,
        source_path: str,
    ) -> Iterator[DocumentNode]:
        fitz = _load_fitz()
        with fitz.open(stream=content, filetype="pdf") as document:
            if document.needs_pass:
                raise ParserError(
                    f"PDF is encrypted and requires a password: {source_path}"
                )
            try:
                yield from iter_pdf_document_nodes(
                    document,
                    document_id=document_id,
                    source_path=source_path,
                )
            except PdfTextLayerError as exc:
                raise ParserError(str(exc)) from exc


class CsvParser(DocumentParser):
    extensions = (".csv",)

    def iter_nodes(
        self,
        path: Path,
        *,
        document_id: str,
        source_path: str,
        encoding: Optional[str] = None,
    ) -> Iterator[DocumentNode]:
        resolved_encoding = _resolve_csv_encoding(
            encoding or _detect_csv_encoding(path, source_path=source_path)
        )
        try:
            with path.open(
                "r",
                encoding=resolved_encoding,
                errors="strict",
                newline="",
            ) as reader:
                yield from _iter_csv_nodes(
                    reader,
                    document_id=document_id,
                    source_path=source_path,
                    encoding=resolved_encoding,
                )
        except UnicodeDecodeError as exc:
            raise CsvEncodingError(
                source_path,
                attempted=(resolved_encoding,),
                selected=resolved_encoding,
            ) from exc

    def iter_nodes_from_bytes(
        self,
        filename: str,
        content: bytes,
        *,
        document_id: str,
        source_path: str,
        encoding: Optional[str] = None,
    ) -> Iterator[DocumentNode]:
        text, resolved_encoding = _decode_csv_bytes(
            content,
            source_path=source_path,
            encoding=encoding,
        )
        yield from _iter_csv_nodes(
            StringIO(text),
            document_id=document_id,
            source_path=source_path,
            encoding=resolved_encoding,
        )


class XlsxParser(DocumentParser):
    extensions = (".xlsx",)

    def iter_nodes(
        self,
        path: Path,
        *,
        document_id: str,
        source_path: str,
    ) -> Iterator[DocumentNode]:
        openpyxl = _load_openpyxl()
        try:
            workbook = openpyxl.load_workbook(
                path,
                read_only=True,
                data_only=False,
            )
        except BadZipFile as exc:
            raise ParserError(
                f"XLSX is damaged or encrypted: {source_path}"
            ) from exc
        sheet_info = _read_xlsx_sheet_info(path, workbook.worksheets)
        cached_workbook = None
        if any(info["has_formulas"] for info in sheet_info.values()):
            cached_workbook = openpyxl.load_workbook(
                path,
                read_only=True,
                data_only=True,
            )
        try:
            yield from _iter_xlsx_nodes(
                workbook,
                document_id=document_id,
                source_path=source_path,
                sheet_info=sheet_info,
                cached_workbook=cached_workbook,
            )
        finally:
            if cached_workbook is not None:
                cached_workbook.close()
            workbook.close()

    def iter_nodes_from_bytes(
        self,
        filename: str,
        content: bytes,
        *,
        document_id: str,
        source_path: str,
    ) -> Iterator[DocumentNode]:
        openpyxl = _load_openpyxl()
        source = BytesIO(content)
        try:
            workbook = openpyxl.load_workbook(
                source,
                read_only=True,
                data_only=False,
            )
        except BadZipFile as exc:
            raise ParserError(
                f"XLSX is damaged or encrypted: {source_path}"
            ) from exc
        sheet_info = _read_xlsx_sheet_info(
            BytesIO(content),
            workbook.worksheets,
        )
        cached_workbook = None
        if any(info["has_formulas"] for info in sheet_info.values()):
            cached_workbook = openpyxl.load_workbook(
                BytesIO(content),
                read_only=True,
                data_only=True,
            )
        try:
            yield from _iter_xlsx_nodes(
                workbook,
                document_id=document_id,
                source_path=source_path,
                sheet_info=sheet_info,
                cached_workbook=cached_workbook,
            )
        finally:
            if cached_workbook is not None:
                cached_workbook.close()
            workbook.close()


_PARSERS: dict[str, DocumentParser] = {}
for _parser in (TextParser(), PdfParser(), CsvParser(), XlsxParser()):
    for _extension in _parser.extensions:
        _PARSERS[_extension] = _parser


def parser_for_path(path: str | Path) -> DocumentParser:
    extension = PurePosixPath(str(path)).suffix.lower()
    try:
        return _PARSERS[extension]
    except KeyError as exc:
        supported = ", ".join(sorted(_PARSERS))
        raise ParserError(
            f"Unsupported document extension {extension!r}. "
            f"Supported extensions: {supported}"
        ) from exc


def iter_document_nodes(
    path: str | Path,
    *,
    document_id: Optional[str] = None,
    source_path: Optional[str] = None,
    csv_encoding: Optional[str] = None,
) -> Iterator[DocumentNode]:
    path = Path(path)
    normalized_source = _normalize_source_path(source_path or path.name)
    resolved_document_id = document_id or document_id_from_source(normalized_source)
    parser = parser_for_path(path)
    if isinstance(parser, CsvParser):
        yield from parser.iter_nodes(
            path,
            document_id=resolved_document_id,
            source_path=normalized_source,
            encoding=csv_encoding,
        )
        return
    yield from parser.iter_nodes(
        path,
        document_id=resolved_document_id,
        source_path=normalized_source,
    )


def iter_document_nodes_from_bytes(
    filename: str,
    content: bytes,
    *,
    document_id: Optional[str] = None,
    source_path: Optional[str] = None,
    csv_encoding: Optional[str] = None,
) -> Iterator[DocumentNode]:
    normalized_source = _normalize_source_path(source_path or filename)
    resolved_document_id = document_id or document_id_from_source(normalized_source)
    parser = parser_for_path(filename)
    if isinstance(parser, CsvParser):
        yield from parser.iter_nodes_from_bytes(
            filename,
            content,
            document_id=resolved_document_id,
            source_path=normalized_source,
            encoding=csv_encoding,
        )
        return
    yield from parser.iter_nodes_from_bytes(
        filename,
        content,
        document_id=resolved_document_id,
        source_path=normalized_source,
    )


def _iter_text_nodes(
    reader: TextIO,
    *,
    document_id: str,
    source_path: str,
    file_type: str,
) -> Iterator[DocumentNode]:
    buffer: list[str] = []
    current_chars = 0
    sequence = 0
    for line in reader:
        buffer.append(line)
        current_chars += len(line)
        if current_chars >= TEXT_NODE_MAX_CHARS:
            text = "".join(buffer).strip()
            if text:
                yield _text_node(
                    text,
                    document_id=document_id,
                    source_path=source_path,
                    file_type=file_type,
                    sequence=sequence,
                )
                sequence += 1
            buffer = []
            current_chars = 0
    text = "".join(buffer).strip()
    if text:
        yield _text_node(
            text,
            document_id=document_id,
            source_path=source_path,
            file_type=file_type,
            sequence=sequence,
        )


def _text_node(
    content: str,
    *,
    document_id: str,
    source_path: str,
    file_type: str,
    sequence: int,
) -> DocumentNode:
    return DocumentNode(
        document_id=document_id,
        content=content,
        parser_version=TEXT_PARSER_VERSION,
        node_type=NodeType.TEXT,
        sequence=sequence,
        source_anchor={
            "source_path": source_path,
            "sequence": sequence,
        },
        metadata={
            "filename": PurePosixPath(source_path).name,
            "file_type": file_type,
        },
    )


def _iter_csv_nodes(
    reader: TextIO,
    *,
    document_id: str,
    source_path: str,
    encoding: str,
) -> Iterator[DocumentNode]:
    sample = reader.read(8192)
    reader.seek(0)
    delimiter = _detect_csv_delimiter(sample)
    csv_reader = csv.reader(reader, delimiter=delimiter)
    try:
        header = next(csv_reader)
    except StopIteration:
        return
    except csv.Error as exc:
        raise ParserError(f"CSV parsing failed for {source_path}: {exc}") from exc
    header = _normalize_headers(header)
    expected_columns = len(header)
    if expected_columns == 0:
        return
    summary = DocumentNode(
        document_id=document_id,
        content=(
            f"CSV 文件：{PurePosixPath(source_path).name}\n"
            f"列名：{'、'.join(header)}"
        ),
        parser_version=CSV_PARSER_VERSION,
        node_type=NodeType.SHEET_SUMMARY,
        sequence=0,
        source_anchor={
            "source_path": source_path,
            "delimiter": delimiter,
            "encoding": encoding,
        },
        metadata={
            "filename": PurePosixPath(source_path).name,
            "file_type": ".csv",
            "columns": header,
        },
    )
    yield summary
    rows: list[list[str]] = []
    row_start = 2
    sequence = 1
    try:
        for row_number, row in enumerate(csv_reader, start=2):
            if len(row) != expected_columns:
                raise CsvStructureError(
                    source_path,
                    row_number=row_number,
                    expected_columns=expected_columns,
                    actual_columns=len(row),
                )
            rows.append([str(value or "") for value in row])
            if len(rows) >= CSV_ROW_GROUP_SIZE:
                yield _row_group_node(
                    rows,
                    header,
                    document_id=document_id,
                    source_path=source_path,
                    parent_id=summary.node_id,
                    sequence=sequence,
                    row_start=row_start,
                    row_end=row_number,
                    row_numbers=list(range(row_start, row_number + 1)),
                    metadata={"delimiter": delimiter, "encoding": encoding},
                )
                sequence += 1
                row_start = row_number + 1
                rows = []
    except csv.Error as exc:
        raise ParserError(f"CSV parsing failed for {source_path}: {exc}") from exc
    if rows:
        yield _row_group_node(
            rows,
            header,
            document_id=document_id,
            source_path=source_path,
            parent_id=summary.node_id,
            sequence=sequence,
            row_start=row_start,
            row_end=row_start + len(rows) - 1,
            row_numbers=list(range(row_start, row_start + len(rows))),
            metadata={"delimiter": delimiter, "encoding": encoding},
        )


def _iter_xlsx_nodes(
    workbook: Any,
    *,
    document_id: str,
    source_path: str,
    sheet_info: dict[str, dict[str, Any]],
    cached_workbook: Any = None,
) -> Iterator[DocumentNode]:
    workbook_summary = _xlsx_workbook_summary_node(
        workbook,
        document_id=document_id,
        source_path=source_path,
    )
    yield workbook_summary
    sequence = 1
    for worksheet in workbook.worksheets:
        cached_worksheet = (
            cached_workbook[worksheet.title]
            if cached_workbook is not None
            else None
        )
        rows = worksheet.iter_rows()
        cached_rows = (
            cached_worksheet.iter_rows()
            if cached_worksheet is not None
            else iter(())
        )
        try:
            first_row = next(rows)
        except StopIteration:
            first_row = ()
        cached_first_row = next(cached_rows, ())
        header_record = _xlsx_row_record(1, first_row, cached_first_row)
        headers = _normalize_headers(header_record["texts"])
        row_records = _iter_xlsx_row_records(rows, cached_rows, start=2)
        first_group: list[dict[str, Any]] = []
        while len(first_group) < XLSX_ROW_GROUP_SIZE:
            try:
                first_group.append(next(row_records))
            except StopIteration:
                break
        column_profile = _analyze_xlsx_columns(headers, first_group)
        info = sheet_info.get(
            worksheet.title,
            {"merged_ranges": [], "has_formulas": False},
        )
        summary_lines = [
            f"Sheet：{worksheet.title}",
            f"列名：{'、'.join(headers)}",
            f"数据范围：{worksheet.max_row} 行 × {worksheet.max_column} 列",
        ]
        if column_profile["key_columns"]:
            summary_lines.append(
                "关键列：" + "、".join(column_profile["key_columns"])
            )
        if column_profile["metric_columns"]:
            summary_lines.append(
                "指标列：" + "、".join(column_profile["metric_columns"])
            )
        summary = DocumentNode(
            document_id=document_id,
            content="\n".join(summary_lines),
            parser_version=XLSX_PARSER_VERSION,
            node_type=NodeType.SHEET_SUMMARY,
            page_or_sheet=worksheet.title,
            sequence=sequence,
            parent_id=workbook_summary.node_id,
            source_anchor={
                "source_path": source_path,
                "sheet": worksheet.title,
                "row_start": 1,
                "row_end": worksheet.max_row,
                "column_start": 1,
                "column_end": worksheet.max_column,
            },
            metadata={
                "filename": PurePosixPath(source_path).name,
                "file_type": ".xlsx",
                "columns": headers,
                "column_types": column_profile["column_types"],
                "key_columns": column_profile["key_columns"],
                "metric_columns": column_profile["metric_columns"],
                "merged_ranges": info["merged_ranges"],
                "has_formulas": info["has_formulas"],
            },
        )
        yield summary
        sequence += 1
        row_group = first_group
        if len(row_group) >= XLSX_ROW_GROUP_SIZE:
            yield _xlsx_row_group_node(
                row_group,
                headers,
                document_id=document_id,
                source_path=source_path,
                parent_id=summary.node_id,
                sequence=sequence,
                sheet=worksheet.title,
                column_types=column_profile["column_types"],
            )
            sequence += 1
            row_group = []
        for record in row_records:
            row_group.append(record)
            if len(row_group) >= XLSX_ROW_GROUP_SIZE:
                yield _xlsx_row_group_node(
                    row_group,
                    headers,
                    document_id=document_id,
                    source_path=source_path,
                    parent_id=summary.node_id,
                    sequence=sequence,
                    sheet=worksheet.title,
                    column_types=column_profile["column_types"],
                )
                sequence += 1
                row_group = []
        if row_group:
            yield _xlsx_row_group_node(
                row_group,
                headers,
                document_id=document_id,
                source_path=source_path,
                parent_id=summary.node_id,
                sequence=sequence,
                sheet=worksheet.title,
                column_types=column_profile["column_types"],
            )


def _xlsx_workbook_summary_node(
    workbook: Any,
    *,
    document_id: str,
    source_path: str,
) -> DocumentNode:
    filename = PurePosixPath(source_path).name
    sheets: list[dict[str, Any]] = []
    lines = [
        f"Excel 文件：{filename}",
        f"Sheet 数量：{len(workbook.worksheets)}",
        "Sheet 列表："
        + "、".join(worksheet.title for worksheet in workbook.worksheets),
    ]
    for worksheet in workbook.worksheets:
        first_row = next(worksheet.iter_rows(min_row=1, max_row=1), ())
        headers = _normalize_headers(
            _xlsx_row_record(1, first_row, ())["texts"]
        )
        sheet = {
            "name": worksheet.title,
            "row_count": int(worksheet.max_row or 0),
            "column_count": int(worksheet.max_column or 0),
            "columns": headers,
        }
        sheets.append(sheet)
        detail = (
            f"{worksheet.title}：{sheet['row_count']} 行 × "
            f"{sheet['column_count']} 列"
        )
        if headers:
            detail += "；列名：" + "、".join(headers)
        lines.append(detail)
    return DocumentNode(
        document_id=document_id,
        content="\n".join(lines),
        parser_version=XLSX_PARSER_VERSION,
        node_type=NodeType.WORKBOOK_SUMMARY,
        sequence=0,
        source_anchor={"source_path": source_path},
        metadata={
            "filename": filename,
            "file_type": ".xlsx",
            "sheet_count": len(sheets),
            "sheets": sheets,
        },
    )


def _row_group_node(
    rows: list[list[str]],
    headers: list[str],
    *,
    document_id: str,
    source_path: str,
    parent_id: str,
    sequence: int,
    row_start: int,
    row_end: int,
    page_or_sheet: Optional[str] = None,
    row_numbers: Optional[list[int]] = None,
    metadata: Optional[dict[str, Any]] = None,
) -> DocumentNode:
    lines = ["\t".join(headers)]
    lines.extend("\t".join(row) for row in rows)
    resolved_row_numbers = row_numbers or list(range(row_start, row_end + 1))
    resolved_metadata = dict(metadata or {})
    resolved_metadata["row_numbers"] = resolved_row_numbers
    return DocumentNode(
        document_id=document_id,
        content="\n".join(lines),
        parser_version=XLSX_PARSER_VERSION if page_or_sheet else CSV_PARSER_VERSION,
        node_type=NodeType.ROW_GROUP,
        page_or_sheet=page_or_sheet,
        parent_id=parent_id,
        sequence=sequence,
        row_start=row_start,
        row_end=row_end,
        column_start=1,
        column_end=max(len(headers), 1),
        source_anchor={
            "source_path": source_path,
            "sheet": page_or_sheet,
            "row_start": row_start,
            "row_end": row_end,
            "column_start": 1,
            "column_end": max(len(headers), 1),
            "row_numbers": resolved_row_numbers,
        },
        metadata=resolved_metadata,
    )


def _xlsx_row_group_node(
    records: list[dict[str, Any]],
    headers: list[str],
    *,
    document_id: str,
    source_path: str,
    parent_id: str,
    sequence: int,
    sheet: str,
    column_types: dict[str, str],
) -> DocumentNode:
    row_numbers = [int(record["row_number"]) for record in records]
    formula_cells = [
        formula
        for record in records
        for formula in record["formula_cells"]
    ]
    metadata: dict[str, Any] = {
        "sheet": sheet,
        "column_types": column_types,
    }
    if formula_cells:
        metadata["formula_cells"] = formula_cells
        metadata["missing_formula_cache_count"] = sum(
            formula["cache_status"] == "missing"
            for formula in formula_cells
        )
    return _row_group_node(
        [record["texts"] for record in records],
        headers,
        document_id=document_id,
        source_path=source_path,
        parent_id=parent_id,
        sequence=sequence,
        row_start=row_numbers[0],
        row_end=row_numbers[-1],
        page_or_sheet=sheet,
        row_numbers=row_numbers,
        metadata=metadata,
    )


def _iter_xlsx_row_records(
    rows: Iterator[tuple[Any, ...]],
    cached_rows: Iterator[tuple[Any, ...]],
    *,
    start: int,
) -> Iterator[dict[str, Any]]:
    for row_number, (row, cached_row) in enumerate(
        zip_longest(rows, cached_rows, fillvalue=()),
        start=start,
    ):
        record = _xlsx_row_record(row_number, row, cached_row)
        if any(record["texts"]) or record["formula_cells"]:
            yield record


def _xlsx_row_record(
    row_number: int,
    row: Iterable[Any],
    cached_row: Iterable[Any],
) -> dict[str, Any]:
    texts: list[str] = []
    kinds: list[str] = []
    formula_cells: list[dict[str, Any]] = []
    for column_number, (cell, cached_cell) in enumerate(
        zip_longest(row, cached_row, fillvalue=None),
        start=1,
    ):
        info = _xlsx_cell_info(cell, cached_cell, row_number, column_number)
        texts.append(info["text"])
        kinds.append(info["kind"])
        if info["formula"] is not None:
            formula_cells.append(info["formula"])
    return {
        "row_number": row_number,
        "texts": texts,
        "kinds": kinds,
        "formula_cells": formula_cells,
    }


def _xlsx_cell_info(
    cell: Any,
    cached_cell: Any,
    row_number: int,
    column_number: int,
) -> dict[str, Any]:
    value = getattr(cell, "value", None)
    data_type = getattr(cell, "data_type", None)
    number_format = str(getattr(cell, "number_format", "") or "General")
    coordinate = getattr(cell, "coordinate", None) or (
        f"R{row_number}C{column_number}"
    )
    if data_type == "f":
        formula_text = str(value or "")
        if formula_text and not formula_text.startswith("="):
            formula_text = "=" + formula_text
        cached_value = getattr(cached_cell, "value", None)
        cached_format = str(
            getattr(cached_cell, "number_format", "") or number_format
        )
        if cached_value is None:
            return {
                "text": f"[公式无缓存值: {formula_text}]",
                "kind": "formula",
                "formula": {
                    "coordinate": coordinate,
                    "formula": formula_text,
                    "cache_status": "missing",
                    "cached_value": None,
                    "number_format": number_format,
                },
            }
        cached_text = _format_excel_value(cached_value, cached_format)
        return {
            "text": f"{cached_text}（公式 {formula_text}）",
            "kind": _excel_value_kind(cached_value, cached_format),
            "formula": {
                "coordinate": coordinate,
                "formula": formula_text,
                "cache_status": "available",
                "cached_value": cached_text,
                "number_format": cached_format,
            },
        }
    return {
        "text": _format_excel_value(value, number_format),
        "kind": _excel_value_kind(value, number_format),
        "formula": None,
    }


def _analyze_xlsx_columns(
    headers: list[str],
    records: list[dict[str, Any]],
) -> dict[str, Any]:
    column_types: dict[str, str] = {}
    key_columns: list[str] = []
    metric_columns: list[str] = []
    for index, header in enumerate(headers):
        kinds = [
            record["kinds"][index]
            for record in records
            if index < len(record["kinds"])
            and record["kinds"][index] != "blank"
        ]
        texts = [
            record["texts"][index]
            for record in records
            if index < len(record["texts"]) and record["texts"][index]
        ]
        if not kinds:
            dominant = "blank"
        else:
            dominant = max(set(kinds), key=lambda item: (kinds.count(item), item))
        column_types[header] = dominant
        if dominant in {"number", "percentage", "currency"}:
            metric_columns.append(header)
        if (
            not key_columns
            and dominant == "text"
            and texts
            and len(set(texts)) / len(texts) >= 0.6
        ):
            key_columns.append(header)
    return {
        "column_types": column_types,
        "key_columns": key_columns,
        "metric_columns": metric_columns,
    }


def _format_excel_value(value: Any, number_format: str) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        format_kind = _number_format_kind(number_format)
        if format_kind == "percentage":
            decimals = _number_format_decimal_places(number_format)
            return f"{float(value) * 100:.{decimals}f}%"
        decimals = _number_format_decimal_places(number_format)
        if decimals > 0:
            return f"{float(value):.{decimals}f}"
        if isinstance(value, int) or float(value).is_integer():
            return str(int(value))
        return format(float(value), ".15g")
    return str(value)


def _excel_value_kind(value: Any, number_format: str) -> str:
    if value is None:
        return "blank"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, time):
        return "time"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return _number_format_kind(number_format)
    return "text"


def _number_format_kind(number_format: str) -> str:
    cleaned = re.sub(r'"[^"]*"', "", str(number_format or ""))
    if "%" in cleaned:
        return "percentage"
    if any(symbol in cleaned for symbol in ("¥", "$", "€", "£", "￥")):
        return "currency"
    return "number"


def _number_format_decimal_places(number_format: str) -> int:
    cleaned = re.sub(r'"[^"]*"', "", str(number_format or ""))
    section = cleaned.split(";", 1)[0]
    match = re.search(r"\.([0#]+)", section)
    return len(match.group(1)) if match else 0


def _read_xlsx_sheet_info(
    source: Any,
    worksheets: Iterable[Any],
) -> dict[str, dict[str, Any]]:
    info: dict[str, dict[str, Any]] = {}
    with ZipFile(source) as archive:
        for worksheet in worksheets:
            merged_ranges: list[str] = []
            has_formulas = False
            worksheet_path = str(
                getattr(worksheet, "_worksheet_path", "")
            ).lstrip("/")
            if worksheet_path:
                with archive.open(worksheet_path) as reader:
                    for _event, element in iterparse(reader, events=("end",)):
                        tag = element.tag.rsplit("}", 1)[-1]
                        if tag == "mergeCell":
                            reference = element.attrib.get("ref")
                            if reference:
                                merged_ranges.append(reference)
                        elif tag == "f":
                            has_formulas = True
                        element.clear()
            info[worksheet.title] = {
                "merged_ranges": merged_ranges,
                "has_formulas": has_formulas,
            }
    return info


def _normalize_headers(values: Iterable[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = str(value or "").strip() or f"column_{index}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def _detect_csv_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        return ","


def _detect_csv_encoding(path: Path, *, source_path: Optional[str] = None) -> str:
    with path.open("rb") as reader:
        prefix = reader.read(4)
    if prefix.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    for encoding in ("utf-8", "gb18030"):
        if _path_decodes(path, encoding):
            return encoding
    raise CsvEncodingError(
        source_path or path.name,
        attempted=CSV_ENCODING_CANDIDATES,
    )


def _decode_csv_bytes(
    content: bytes,
    *,
    source_path: str,
    encoding: Optional[str] = None,
) -> tuple[str, str]:
    if encoding:
        resolved = _resolve_csv_encoding(encoding)
        try:
            return content.decode(resolved, errors="strict"), resolved
        except UnicodeDecodeError as exc:
            raise CsvEncodingError(
                source_path,
                attempted=(resolved,),
                selected=resolved,
            ) from exc
    if content.startswith(b"\xef\xbb\xbf"):
        return content.decode("utf-8-sig"), "utf-8-sig"
    for candidate in ("utf-8", "gb18030"):
        try:
            return content.decode(candidate, errors="strict"), candidate
        except UnicodeDecodeError:
            continue
    raise CsvEncodingError(
        source_path,
        attempted=CSV_ENCODING_CANDIDATES,
    )


def _path_decodes(path: Path, encoding: str) -> bool:
    decoder = codecs.getincrementaldecoder(encoding)(errors="strict")
    try:
        with path.open("rb") as reader:
            while True:
                chunk = reader.read(1024 * 1024)
                if not chunk:
                    break
                decoder.decode(chunk)
        decoder.decode(b"", final=True)
        return True
    except UnicodeDecodeError:
        return False


def _resolve_csv_encoding(encoding: str) -> str:
    try:
        return codecs.lookup(str(encoding).strip()).name
    except LookupError as exc:
        raise ParserError(f"Unknown CSV encoding: {encoding!r}.") from exc


def _normalize_source_path(value: str) -> str:
    parts = [
        part
        for part in PurePosixPath(str(value).replace("\\", "/")).parts
        if part not in ("", ".", "..")
    ]
    return PurePosixPath(*parts).as_posix() if parts else "untitled"


def _load_fitz() -> Any:
    try:
        import fitz
    except ImportError as exc:
        raise ParserDependencyError(
            "PDF parsing requires PyMuPDF. Install requirements.txt first."
        ) from exc
    return fitz


def _load_openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as exc:
        raise ParserDependencyError(
            "XLSX parsing requires openpyxl. Install requirements.txt first."
        ) from exc
    return openpyxl
